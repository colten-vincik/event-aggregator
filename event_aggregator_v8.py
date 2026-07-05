"""
Event & Activities Aggregator v8
==================================
US National Coverage — parallel fetching + expanded sources

NEW in v8:
  - Parallel fetching via ThreadPoolExecutor (cities run concurrently,
    sources within each city run concurrently)
  - Per-domain rate limiter (polite gaps enforced per hostname, not globally)
  - DoStuff network: 24 city sites (do312, do617, do214, do215, do512...)
  - Bandsintown concert listings by city
  - Yelp Events city pages

SOURCES:
  Universal (45+ cities):
    TimeOut · Patch.com · Eventbrite · Songkick · AllEvents.in
    DoStuff network · Bandsintown · Yelp Events

  NYC-specific:
    NYC Open Data Permitted Events (2000)
    NYC Open Data Film & Theatre Permits (200)
    SummerStage · Prospect Park

  Attractions (coordinate-based):
    Wikidata SPARQL · OpenStreetMap Overpass

  Attractions (NYC-specific):
    NYC Cultural Organizations · NYC Farmers Markets

Requirements:
    pip install requests beautifulsoup4 lxml openpyxl
"""

import re, sys, time, json, random, threading, math
from datetime import datetime
from urllib.parse import quote_plus, urlparse
from collections import Counter
from functools import partial
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

# Load .env if present (local dev)
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Run logger ───────────────────────────────────────────────────────────────

class RunLog:
    """
    Thread-safe run logger.  Tracks every tprint line, every HTTP failure,
    and a per-source result entry.  Call write() at the end of a run to
    produce a human-readable report file.
    """
    def __init__(self):
        self._lock   = threading.Lock()
        self.lines   = []   # all console output lines with timestamps
        self.errors  = []   # HTTP-level failures: (ts, url, error_str)
        self.sources = []   # per-source results dicts

    def line(self, msg: str):
        with self._lock:
            ts = datetime.now().strftime("%H:%M:%S")
            self.lines.append(f"[{ts}] {msg}")

    def http_error(self, url: str, error):
        with self._lock:
            ts = datetime.now().strftime("%H:%M:%S")
            self.errors.append((ts, url, str(error)))

    def source(self, name: str, city: str, count: int, error: str = ""):
        with self._lock:
            self.sources.append({
                "source": name,
                "city":   city,
                "count":  count,
                "ok":     not error,
                "error":  error,
            })

    def write(self, path: str, label: str, elapsed: float):
        W = 82
        sep = "─" * W
        out = []

        out.append("=" * W)
        out.append("  Event & Activities Aggregator v8 — Run Report")
        out.append("=" * W)
        out.append(f"  Search:    {label}")
        out.append(f"  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        out.append(f"  Duration:  {elapsed:.0f}s")
        out.append("")

        # ── Source results ────────────────────────────────────────────────────
        ok_count   = sum(1 for s in self.sources if s["ok"])
        fail_count = sum(1 for s in self.sources if not s["ok"])
        zero_count = sum(1 for s in self.sources if s["ok"] and s["count"] == 0)
        out.append(f"SOURCE RESULTS  ({ok_count} OK, {fail_count} failed, {zero_count} returned 0 results)")
        out.append(sep)
        out.append(f"  {'Status':<8} {'Source':<22} {'City':<24} {'Results':>7}  Notes")
        out.append(sep)
        for s in sorted(self.sources, key=lambda x: (x["city"], x["source"])):
            status = "✓ OK  " if s["ok"] else "✗ FAIL"
            count  = str(s["count"]) if s["ok"] else "—"
            note   = s["error"][:45] if s["error"] else ("(no results)" if s["count"] == 0 else "")
            out.append(f"  {status}  {s['source']:<22} {s['city']:<24} {count:>7}  {note}")
        out.append("")

        # ── HTTP errors ───────────────────────────────────────────────────────
        if self.errors:
            out.append(f"HTTP ERRORS  ({len(self.errors)} total)")
            out.append(sep)
            for ts, url, err in self.errors:
                out.append(f"  [{ts}] {url[:72]}")
                out.append(f"         → {err[:74]}")
                out.append("")
        else:
            out.append("HTTP ERRORS  (none)")
            out.append(sep)
            out.append("  All requests succeeded at the network level.")
            out.append("")

        # ── Full console log ──────────────────────────────────────────────────
        out.append(f"FULL CONSOLE LOG  ({len(self.lines)} lines)")
        out.append(sep)
        out.extend(self.lines)

        with open(path, "w", encoding="utf-8") as f:
            f.write("\n".join(out) + "\n")
        print(f"  📄  Report saved: {path}")

RUNLOG = RunLog()

# ─── Per-domain rate limiter ──────────────────────────────────────────────────

class DomainLimiter:
    """
    Serializes requests to each hostname with a configurable minimum gap.
    Multiple threads can hit different domains simultaneously, but requests
    to the same domain are queued and spaced out.
    """
    def __init__(self):
        self._meta  = threading.Lock()
        self._slots = {}  # hostname -> {"lock": Lock, "last": float}

    def _slot(self, host):
        with self._meta:
            if host not in self._slots:
                self._slots[host] = {"lock": threading.Lock(), "last": 0.0}
            return self._slots[host]

    def wait(self, url, min_gap=1.1):
        host = urlparse(url).netloc
        slot = self._slot(host)
        with slot["lock"]:
            elapsed = time.time() - slot["last"]
            gap = min_gap + random.uniform(0, 0.3)
            if elapsed < gap:
                time.sleep(gap - elapsed)
            slot["last"] = time.time()

LIMITER = DomainLimiter()
_PRINT_LOCK = threading.Lock()

def tprint(*args, **kwargs):
    msg = " ".join(str(a) for a in args)
    with _PRINT_LOCK:
        print(msg)
        RUNLOG.line(msg)

# ─── HTTP ─────────────────────────────────────────────────────────────────────

SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,*/*;q=0.8",
})

def get(url, min_gap=1.1, hdrs=None, timeout=15, silent_404=False):
    LIMITER.wait(url, min_gap)
    try:
        r = SESSION.get(url, headers=hdrs or {}, timeout=timeout)
        r.raise_for_status()
        return r
    except requests.HTTPError as e:
        code = e.response.status_code if e.response is not None else 0
        if silent_404 and code == 404:
            return None  # expected miss — don't log
        tprint(f"  ⚠  GET {code} {url[:68]}…")
        RUNLOG.http_error(url, e)
        return None
    except Exception as e:
        tprint(f"  ⚠  GET  {url[:70]}… → {e}")
        RUNLOG.http_error(url, e)
        return None

def post_req(url, data, min_gap=1.5, hdrs=None, timeout=25):
    LIMITER.wait(url, min_gap)
    try:
        r = SESSION.post(url, data=data, headers=hdrs or {}, timeout=timeout)
        r.raise_for_status()
        return r
    except Exception as e:
        tprint(f"  ⚠  POST {url[:70]}… → {e}")
        RUNLOG.http_error(url, e)
        return None

# ─── Styling ──────────────────────────────────────────────────────────────────

CAT_COLOURS = {
    "Music":         "D6E4F7",
    "Sports":        "D5E8D4",
    "Arts & Theatre":"E1D5E7",
    "Food & Drink":  "FFE6CC",
    "Community":     "FFF2CC",
    "Nature & Parks":"C9E6C9",
    "Museums":       "E8DAEF",
    "Amusement":     "FDDCB5",
    "Shopping":      "FEF9E7",
    "Entertainment": "FDE9D9",
    "Architecture":  "F8CECC",
    "Film":          "DAE8FC",
    "Education":     "E8DAEF",
    "Other":         "F5F5F5",
}

def _s(): return Side(style="thin", color="CCCCCC")
def _b(): return Border(left=_s(), right=_s(), top=_s(), bottom=_s())

def hdr(cell, bg="1F3864", fg="FFFFFF", sz=11):
    cell.font      = Font(bold=True, color=fg, name="Arial", size=sz)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _b()

def dat(cell, bg="FFFFFF"):
    cell.font      = Font(name="Arial", size=10)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border    = _b()

def set_widths(ws, d):
    for col, w in d.items():
        ws.column_dimensions[col].width = w

# ─── City Configuration ───────────────────────────────────────────────────────
# Keys per city:
#   state    — two-letter state abbreviation
#   timeout  — TimeOut slug (None if not covered)
#   eb       — Eventbrite city URL segment
#   sk       — Songkick metro area integer ID (None if unknown)
#   patch    — list of (state-slug, city-slug) for Patch.com
#   dostuff  — DoStuff network domain (None if not covered)

CITY_CONFIG = {
    # sk_slug: full verified Songkick metro-area slug from songkick.com/metro-areas/{sk_slug}
    #          None = unverified, source will be skipped cleanly rather than 406-ing
    #
    # ── Northeast ─────────────────────────────────────────────────────────────
    "new york":       {"state":"NY","timeout":"newyork",       "eb":"ny--new-york-city",  "sk_slug":"7644-us-new-york-nyc",          "patch":[("new-york","new-york-city"),("new-york","upper-west-side-nyc"),("new-york","brooklyn"),("new-york","queens")], "dostuff":None},
    "boston":         {"state":"MA","timeout":"boston",        "eb":"ma--boston",         "sk_slug":"18842-us-boston-cambridge",     "patch":[("massachusetts","boston")],                                                                                      "dostuff":"do617.com"},
    "philadelphia":   {"state":"PA","timeout":"philadelphia",  "eb":"pa--philadelphia",   "sk_slug":"5202-us-philadelphia",          "patch":[("pennsylvania","philadelphia")],                                                                                  "dostuff":"do215.com"},
    "washington":     {"state":"DC","timeout":"washington-dc", "eb":"dc--washington",     "sk_slug":"1409-us-washington",            "patch":[("dc","washington")],                                                                                             "dostuff":"do202.com"},
    "baltimore":      {"state":"MD","timeout":"baltimore",     "eb":"md--baltimore",      "sk_slug":"4125-us-baltimore",             "patch":[("maryland","baltimore")],                                                                                        "dostuff":"do410.com"},
    "pittsburgh":     {"state":"PA","timeout":"pittsburgh",    "eb":"pa--pittsburgh",     "sk_slug":None,                           "patch":[("pennsylvania","pittsburgh")],                                                                                    "dostuff":"do412.com"},
    "buffalo":        {"state":"NY","timeout":None,            "eb":"ny--buffalo",        "sk_slug":None,                           "patch":[("new-york","buffalo")],                                                                                           "dostuff":None},
    "hartford":       {"state":"CT","timeout":None,            "eb":"ct--hartford",       "sk_slug":None,                           "patch":[("connecticut","hartford")],                                                                                       "dostuff":None},
    "providence":     {"state":"RI","timeout":None,            "eb":"ri--providence",     "sk_slug":None,                           "patch":[("rhode-island","providence")],                                                                                    "dostuff":None},
    "albany":         {"state":"NY","timeout":None,            "eb":"ny--albany",         "sk_slug":None,                           "patch":[("new-york","albany")],                                                                                            "dostuff":None},
    "new haven":      {"state":"CT","timeout":None,            "eb":"ct--new-haven",      "sk_slug":None,                           "patch":[("connecticut","new-haven")],                                                                                      "dostuff":None},
    # ── Southeast ─────────────────────────────────────────────────────────────
    "miami":          {"state":"FL","timeout":"miami",         "eb":"fl--miami",          "sk_slug":"9776-us-miami",                 "patch":[("florida","miami")],                                                                                              "dostuff":"do305.com"},
    "atlanta":        {"state":"GA","timeout":"atlanta",       "eb":"ga--atlanta",        "sk_slug":"4120-us-atlanta",               "patch":[("georgia","atlanta")],                                                                                            "dostuff":"do404.com"},
    "nashville":      {"state":"TN","timeout":"nashville",     "eb":"tn--nashville",      "sk_slug":"11104-us-nashville",            "patch":[("tennessee","nashville")],                                                                                        "dostuff":"do615.com"},
    "charlotte":      {"state":"NC","timeout":"charlotte",     "eb":"nc--charlotte",      "sk_slug":None,                           "patch":[("north-carolina","charlotte")],                                                                                    "dostuff":"do704.com"},
    "orlando":        {"state":"FL","timeout":"orlando",       "eb":"fl--orlando",        "sk_slug":None,                           "patch":[("florida","orlando")],                                                                                             "dostuff":None},
    "tampa":          {"state":"FL","timeout":"tampa",         "eb":"fl--tampa",          "sk_slug":None,                           "patch":[("florida","tampa")],                                                                                               "dostuff":None},
    "new orleans":    {"state":"LA","timeout":"new-orleans",   "eb":"la--new-orleans",    "sk_slug":None,                           "patch":[("louisiana","new-orleans")],                                                                                       "dostuff":"do504.com"},
    "raleigh":        {"state":"NC","timeout":"raleigh",       "eb":"nc--raleigh",        "sk_slug":None,                           "patch":[("north-carolina","raleigh")],                                                                                      "dostuff":"do919.com"},
    "richmond":       {"state":"VA","timeout":None,            "eb":"va--richmond",       "sk_slug":None,                           "patch":[("virginia","richmond")],                                                                                          "dostuff":"do804.com"},
    "memphis":        {"state":"TN","timeout":"memphis",       "eb":"tn--memphis",        "sk_slug":None,                           "patch":[("tennessee","memphis")],                                                                                          "dostuff":None},
    "louisville":     {"state":"KY","timeout":None,            "eb":"ky--louisville",     "sk_slug":None,                           "patch":[("kentucky","louisville")],                                                                                        "dostuff":None},
    "jacksonville":   {"state":"FL","timeout":None,            "eb":"fl--jacksonville",   "sk_slug":None,                           "patch":[("florida","jacksonville")],                                                                                       "dostuff":None},
    "savannah":       {"state":"GA","timeout":None,            "eb":"ga--savannah",       "sk_slug":None,                           "patch":[("georgia","savannah")],                                                                                           "dostuff":None},
    # ── Midwest ───────────────────────────────────────────────────────────────
    "chicago":        {"state":"IL","timeout":"chicago",       "eb":"il--chicago",        "sk_slug":"9426-us-chicago",               "patch":[("illinois","chicago")],                                                                                           "dostuff":"do312.com"},
    "detroit":        {"state":"MI","timeout":"detroit",       "eb":"mi--detroit",        "sk_slug":None,                           "patch":[("michigan","detroit")],                                                                                            "dostuff":None},
    "minneapolis":    {"state":"MN","timeout":"minneapolis",   "eb":"mn--minneapolis",    "sk_slug":None,                           "patch":[("minnesota","minneapolis")],                                                                                       "dostuff":"do612.com"},
    "columbus":       {"state":"OH","timeout":"columbus",      "eb":"oh--columbus",       "sk_slug":None,                           "patch":[("ohio","columbus")],                                                                                               "dostuff":"do614.com"},
    "indianapolis":   {"state":"IN","timeout":"indianapolis",  "eb":"in--indianapolis",   "sk_slug":None,                           "patch":[("indiana","indianapolis")],                                                                                        "dostuff":"do317.com"},
    "kansas city":    {"state":"MO","timeout":"kansas-city",   "eb":"mo--kansas-city",    "sk_slug":None,                           "patch":[("missouri","kansas-city")],                                                                                        "dostuff":"do816.com"},
    "st louis":       {"state":"MO","timeout":"st-louis",      "eb":"mo--st-louis",       "sk_slug":None,                           "patch":[("missouri","st-louis")],                                                                                          "dostuff":"do314.com"},
    "cincinnati":     {"state":"OH","timeout":"cincinnati",    "eb":"oh--cincinnati",     "sk_slug":None,                           "patch":[("ohio","cincinnati")],                                                                                             "dostuff":None},
    "cleveland":      {"state":"OH","timeout":"cleveland",     "eb":"oh--cleveland",      "sk_slug":None,                           "patch":[("ohio","cleveland")],                                                                                              "dostuff":"do216.com"},
    "milwaukee":      {"state":"WI","timeout":"milwaukee",     "eb":"wi--milwaukee",      "sk_slug":None,                           "patch":[("wisconsin","milwaukee")],                                                                                         "dostuff":None},
    # ── Southwest / Mountain ──────────────────────────────────────────────────
    "dallas":         {"state":"TX","timeout":"dallas",        "eb":"tx--dallas",         "sk_slug":None,                           "patch":[("texas","dallas")],                                                                                                "dostuff":"do214.com"},
    "houston":        {"state":"TX","timeout":"houston",       "eb":"tx--houston",        "sk_slug":"15073-us-houston",              "patch":[("texas","houston")],                                                                                               "dostuff":"do713.com"},
    "austin":         {"state":"TX","timeout":"austin",        "eb":"tx--austin",         "sk_slug":"9179-us-austin",                "patch":[("texas","austin")],                                                                                                "dostuff":"do512.com"},
    "san antonio":    {"state":"TX","timeout":None,            "eb":"tx--san-antonio",    "sk_slug":None,                           "patch":[("texas","san-antonio")],                                                                                          "dostuff":"do210.com"},
    "phoenix":        {"state":"AZ","timeout":"phoenix",       "eb":"az--phoenix",        "sk_slug":None,                           "patch":[("arizona","phoenix")],                                                                                             "dostuff":None},
    "denver":         {"state":"CO","timeout":"denver",        "eb":"co--denver",         "sk_slug":"6404-us-denver",                "patch":[("colorado","denver")],                                                                                             "dostuff":"do303.com"},
    "las vegas":      {"state":"NV","timeout":"las-vegas",     "eb":"nv--las-vegas",      "sk_slug":"8396-us-las-vegas",             "patch":[("nevada","las-vegas")],                                                                                            "dostuff":None},
    "albuquerque":    {"state":"NM","timeout":None,            "eb":"nm--albuquerque",    "sk_slug":None,                           "patch":[("new-mexico","albuquerque")],                                                                                      "dostuff":None},
    "tucson":         {"state":"AZ","timeout":None,            "eb":"az--tucson",         "sk_slug":None,                           "patch":[("arizona","tucson")],                                                                                              "dostuff":None},
    "oklahoma city":  {"state":"OK","timeout":None,            "eb":"ok--oklahoma-city",  "sk_slug":None,                           "patch":[("oklahoma","oklahoma-city")],                                                                                      "dostuff":None},
    # ── West Coast / Pacific ──────────────────────────────────────────────────
    "los angeles":    {"state":"CA","timeout":"los-angeles",   "eb":"ca--los-angeles",    "sk_slug":"17835-us-los-angeles-la",       "patch":[("california","los-angeles")],                                                                                      "dostuff":"do213.com"},
    "san francisco":  {"state":"CA","timeout":"san-francisco", "eb":"ca--san-francisco",  "sk_slug":"26330-us-sf-bay-area",          "patch":[("california","san-francisco")],                                                                                    "dostuff":"do415.com"},
    "seattle":        {"state":"WA","timeout":"seattle",       "eb":"wa--seattle",        "sk_slug":"2846-us-seattle",               "patch":[("washington","seattle")],                                                                                          "dostuff":"do206.com"},
    "portland":       {"state":"OR","timeout":"portland",      "eb":"or--portland",       "sk_slug":None,                           "patch":[("oregon","portland")],                                                                                             "dostuff":"do503.com"},
    "san diego":      {"state":"CA","timeout":"san-diego",     "eb":"ca--san-diego",      "sk_slug":"11086-us-san-diego",            "patch":[("california","san-diego")],                                                                                        "dostuff":None},
    "sacramento":     {"state":"CA","timeout":None,            "eb":"ca--sacramento",     "sk_slug":None,                           "patch":[("california","sacramento")],                                                                                       "dostuff":None},
    "salt lake city": {"state":"UT","timeout":"salt-lake-city","eb":"ut--salt-lake-city", "sk_slug":None,                           "patch":[("utah","salt-lake-city")],                                                                                         "dostuff":None},
    "boise":          {"state":"ID","timeout":None,            "eb":"id--boise",          "sk_slug":None,                           "patch":[("idaho","boise")],                                                                                                 "dostuff":None},
}

REGIONS = {
    "northeast": ["new york","boston","philadelphia","washington","baltimore","pittsburgh","buffalo","hartford","providence","albany","new haven"],
    "southeast": ["miami","atlanta","nashville","charlotte","orlando","tampa","new orleans","raleigh","richmond","memphis","louisville","jacksonville","savannah"],
    "midwest":   ["chicago","detroit","minneapolis","columbus","indianapolis","kansas city","st louis","cincinnati","cleveland","milwaukee"],
    "southwest": ["dallas","houston","austin","san antonio","phoenix","denver","las vegas","albuquerque","tucson","oklahoma city"],
    "west":      ["los angeles","san francisco","seattle","portland","san diego","sacramento","salt lake city","boise"],
}

def _city_key(location):
    city = location.split(",")[0].strip().lower()
    if city in CITY_CONFIG: return city
    for k in CITY_CONFIG:
        if k in city or city in k: return k
    return None

def _bit_slug(city_key, cfg):
    """Bandsintown / Yelp Events slug:  chicago-il"""
    return f"{city_key.replace(' ','-')}-{cfg.get('state','').lower()}"

# ─── Helpers ──────────────────────────────────────────────────────────────────

def classify(text):
    t = (text or "").lower()
    if any(w in t for w in ["concert","music","band","dj","jazz","rock","hip hop","singer","tour","symphony","orchestra","opera","festival","gig","bluegrass","electronic","choir","r&b","reggae","punk","indie","album","live music","nightlife","party","club","lounge","rave","edm","country music","folk music","open mic"]):
        return "Music"
    if any(w in t for w in ["sport","game","match","nba","nfl","mlb","nhl","soccer","tennis","golf","marathon","race","baseball","basketball","football","hockey","esport","wrestling","mma","athletic","pickleball","lacrosse","volleyball","swim","yoga","fitness","run","cycling","triathlon","5k","10k"]):
        return "Sports"
    if any(w in t for w in ["museum","exhibit","gallery","art show","theatre","theater","ballet","dance","comedy","stand-up","broadway","performance","circus","puppet","magic","improv","opera","cabaret","burlesque","sketch","drag"]):
        return "Arts & Theatre"
    if any(w in t for w in ["food","drink","wine","beer","brunch","tasting","chef","culinary","cocktail","dining","grill","whiskey","spirits","gastro","restaurant","greenmarket","farmers market","food festival","bbq","taco","pizza","coffee","brewery","winery","mixology","distillery"]):
        return "Food & Drink"
    if any(w in t for w in ["community","fair","parade","volunteer","charity","fundraiser","seminar","networking","conference","health","wellness","meditation","block party","street fair","flea","craft","expo","summit","meetup","social","mixer"]):
        return "Community"
    if any(w in t for w in ["park","garden","nature","hike","trail","zoo","aquarium","botanical","outdoor","beach","lake","forest","reserve","greenway","conservancy","wildlife","kayak","canoe","paddl"]):
        return "Nature & Parks"
    if any(w in t for w in ["theme park","amusement","roller coaster","arcade","carnival","carousel","ferris wheel"]):
        return "Amusement"
    if any(w in t for w in ["film","cinema","movie","screening","documentary","short film"]):
        return "Film"
    if any(w in t for w in ["lecture","learn","education","school","university","training","class","workshop","tutorial","course"]):
        return "Education"
    if any(w in t for w in ["shop","mall","flea market","boutique","bazaar","retail","store","market","pop-up"]):
        return "Shopping"
    return "Entertainment"

def parse_date(s):
    if not s: return ""
    s = str(s)
    m = re.search(r"(\d{4}-\d{2}-\d{2})", s)
    if m: return m.group(1)
    if re.match(r"^\d{10}$", s.strip()):
        try: return datetime.fromtimestamp(int(s)).strftime("%Y-%m-%d")
        except: pass
    for fmt in ("%B %d, %Y","%b %d, %Y","%m/%d/%Y","%B %d","%b %d"):
        try:
            dt = datetime.strptime(s.strip()[:20], fmt)
            if dt.year == 1900: dt = dt.replace(year=datetime.today().year)
            return dt.strftime("%Y-%m-%d")
        except: pass
    return s[:20]

def parse_time(s):
    if not s: return ""
    m = re.search(r"T(\d{2}:\d{2})", str(s))
    if m: return m.group(1)
    m = re.search(r"(\d{1,2}:\d{2}\s*(?:am|pm)?)", str(s), re.I)
    if m: return m.group(1)
    return ""

_GEO_LOCK  = threading.Lock()
_GEO_CACHE = {}

def geocode(location):
    with _GEO_LOCK:
        if location in _GEO_CACHE:
            return _GEO_CACHE[location]
    url = (f"https://nominatim.openstreetmap.org/search"
           f"?q={quote_plus(location)}&format=json&limit=1&countrycodes=us")
    r = get(url, min_gap=1.0, hdrs={"User-Agent":"EventAggregator/8.0"})
    if not r:
        with _GEO_LOCK: _GEO_CACHE[location] = (None, None, location)
        return None, None, location
    data = r.json()
    if not data:
        with _GEO_LOCK: _GEO_CACHE[location] = (None, None, location)
        return None, None, location
    d = data[0]
    result = float(d["lat"]), float(d["lon"]), d["display_name"]
    with _GEO_LOCK: _GEO_CACHE[location] = result
    return result

def ev(name, cat, date="", time_="", venue="", address="", city="", price="", url="", source=""):
    return {"Name":name,"Category":cat,"Date":date,"Time":time_,
            "Venue":venue,"Address":address,"City":city,
            "Price":price,"URL":url,"Source":source}

def att(name, cat, type_="", address="", city="", url="", source="",
        lat=None, lon=None, desc=""):
    return {"Name":name,"Category":cat,"Type":type_,"Description":desc,
            "Address":address,"City":city,"URL":url,"Source":source,
            "lat":lat,"lon":lon}

# Todays's date string (YYYY-MM-DD) used for future-date filtering
_TODAY_STR = datetime.today().strftime("%Y-%m-%d")

# List-article patterns — event names matching these are aggregator links, not real events
_LIST_RE = re.compile(
    r"(?i)^(\d+\s+)?(best|top|things?\s+to\s+do|what('?s)?\s+(on|to\s+do|happening)|"
    r"where\s+to|guide\s+to|weekend\s+guide|events?\s+(this|in|near|for|calendar)|"
    r"upcoming\s+events?|featured\s+events?|concert\s+(schedule|calendar|picks?)|"
    r"movies?\s+(now\s+)?showing|what\s+to\s+do|(\w+\s+)?picks?\s+(for|this)|"
    r"things?\s+(to\s+see|happening)|highlights?)"
)

def _is_list_article(name: str) -> bool:
    return bool(_LIST_RE.search(name or ""))

# URLs that point to songs/albums/streaming rather than event pages
_NON_EVENT_URL_RE = re.compile(
    r"(spotify\.com|apple\.com/music|music\.apple\.com|soundcloud\.com"
    r"|bandcamp\.com|tidal\.com|deezer\.com|pandora\.com"
    r"|youtube\.com/watch\?v=|/songs?/|/albums?/|/tracks?/)",
    re.I
)

def _is_valid_event(e: dict) -> bool:
    """Gate: True only when e is a real, upcoming, linkable event.

    Rules enforced:
      • Name must be at least 3 characters and not a list-article title
      • URL must be present and must not point to a music-streaming or song/album page
      • Date must be present
      • If date is a parseable YYYY-MM-DD it must be today or in the future
    """
    name = (e.get("Name") or "").strip()
    url  = (e.get("URL")  or "").strip()
    date = (e.get("Date") or "").strip()

    if len(name) < 3:                    return False
    if _is_list_article(name):           return False
    if not url:                          return False
    if _NON_EVENT_URL_RE.search(url):    return False
    if not date:                         return False
    # Only reject when we have a clean ISO date and it's clearly in the past
    if re.match(r"^\d{4}-\d{2}-\d{2}$", date) and date < _TODAY_STR:
        return False
    return True

def _ld_events(soup, city_label, source_label):
    """Extract schema.org Event items from all JSON-LD blocks on a page."""
    out = []
    event_types = {"Event","MusicEvent","SportsEvent","TheaterEvent",
                   "FoodEvent","SocialEvent","Festival","ExhibitionEvent","ComedyEvent"}
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            d = json.loads(script.string or "")
            items = d if isinstance(d, list) else [d]
            for item in items:
                if item.get("@type") not in event_types: continue
                name = (item.get("name") or "").strip()
                if not name: continue
                loc = item.get("location", {})
                venue = loc.get("name","") if isinstance(loc, dict) else str(loc)
                addr_obj = (loc.get("address",{}) if isinstance(loc, dict) else {})
                address = (addr_obj.get("streetAddress","") if isinstance(addr_obj, dict) else str(addr_obj))
                offers = item.get("offers", {})
                if isinstance(offers, list): offers = offers[0] if offers else {}
                price_val = offers.get("price","") if isinstance(offers, dict) else ""
                price = "Free" if str(price_val) in ("0","0.0") else (f"${price_val}" if price_val else "")
                start = item.get("startDate","")
                out.append(ev(name, classify((item.get("description","") or "")+" "+name),
                    parse_date(start), parse_time(start),
                    venue, address, city_label, price,
                    item.get("url",""), source_label))
        except: pass
    return out

# ═══════════════════════════════════════════════════════════════════════════════
#  EVENT SOURCES — UNIVERSAL
# ═══════════════════════════════════════════════════════════════════════════════

# ── 1. TimeOut ────────────────────────────────────────────────────────────────

TIMEOUT_PAGES = [
    ("music","Music"),("theater","Arts & Theatre"),("art","Arts & Theatre"),
    ("comedy","Arts & Theatre"),("film","Film"),("nightlife","Entertainment"),
    ("dance","Arts & Theatre"),("restaurants","Food & Drink"),
    ("shopping","Shopping"),("things-to-do","Entertainment"),("attractions","Entertainment"),
]
TIMEOUT_SUBPAGES_NYC = [
    ("music/concert-schedule-for-live-music-in-new-york-city","Music"),
    ("nightlife/best-parties-in-nyc-this-week","Entertainment"),
    ("dance/the-best-dance-shows-in-nyc-this-month","Arts & Theatre"),
]

# URL path patterns that indicate a TimeOut list article rather than an individual event page.
# These pages need to be followed to extract the individual items inside.
_TIMEOUT_LIST_RE = re.compile(
    r"/(best-|top-\d|things-to-do|guide-to|what-to-do|where-to|"
    r"this-week|this-weekend|this-month|this-summer|this-fall|"
    r"to-see-|to-do-|-right-now|schedule-for|what-s-on|whats-on)",
    re.I
)
# Non-article TimeOut URL path segments — these look like individual pages
_TIMEOUT_SKIP_HREF_RE = re.compile(
    r"(facebook|twitter|instagram|tiktok|youtube|timeout\.com/news"
    r"|timeout\.com/book|/sponsored|/advertise|mailto:|#)",
    re.I
)

def _timeout_cat_from_href(href, default):
    for tok, mapped in [
        ("music","Music"),("theater","Arts & Theatre"),("art","Arts & Theatre"),
        ("comedy","Arts & Theatre"),("film","Film"),("nightlife","Entertainment"),
        ("dance","Arts & Theatre"),("restaurant","Food & Drink"),
        ("shopping","Shopping"),("attraction","Entertainment"),
    ]:
        if tok in href: return mapped
    return default

def _timeout_follow_list(url, cat, seen, location, base="https://www.timeout.com"):
    """Follow a TimeOut list-article URL and extract the individual named items within it.
    Uses a shorter timeout (8 s) and smaller gap (0.7 s) since the caller already
    throttled the initial listing page fetch. Returns a list of ev() dicts."""
    r = get(url, min_gap=0.7, timeout=8, silent_404=True)
    if not r: return []
    soup = BeautifulSoup(r.text, "lxml")

    # Try JSON-LD first — the richest source when available
    ld = _ld_events(soup, location, "TimeOut")
    if ld:
        return [x for x in ld if x.get("Name") and not _is_list_article(x["Name"])]

    results = []
    # TimeOut list articles mark each entry as a numbered heading near a link.
    # Pattern: <h3>1. Show Name</h3> with an <a> nearby, or <a><h3>1. Show Name</h3></a>
    for tag_el in soup.select("h2, h3, h4"):
        raw = tag_el.get_text(strip=True)
        # Must look like "3. Something" or just have meaningful text
        title = re.sub(r"^\d+[\.\)]\s*", "", raw).strip()
        if not title or len(title) < 4 or _is_list_article(title): continue
        key = title.lower()[:60]
        if key in seen: continue
        # Find the link: either the heading is inside <a>, or the next/parent <a>
        a = tag_el.find("a", href=True)
        if not a:
            a = tag_el.find_parent("a", href=True)
        if not a:
            # Look at the next sibling element for a link
            sib = tag_el.find_next_sibling()
            if sib: a = sib.find("a", href=True)
        href = ""
        if a:
            href = a.get("href","")
            if href and not href.startswith("http"): href = base + href
        if href and _TIMEOUT_SKIP_HREF_RE.search(href): continue
        seen.add(key)
        results.append(ev(title, _timeout_cat_from_href(href, cat),
                          city=location, url=href or url, source="TimeOut"))
    return results

def _timeout_cards(soup, default_cat, seen, base="https://www.timeout.com"):
    """Scrape card tiles from a TimeOut listing page.
    Returns (title, cat, href, is_list_article) tuples."""
    added = []
    processed = set()
    for card in soup.select("article[data-testid='tile-zone-a_testID'], article"):
        link = card.find("a", href=True)
        if not link: continue
        href = link["href"]
        if not href.startswith("http"): href = base + href
        if href in processed or _TIMEOUT_SKIP_HREF_RE.search(href): continue
        processed.add(href)
        img   = card.find("img")
        title = (img.get("alt","") if img else "").strip()
        if not title:
            h = card.find(["h2","h3","h4"])
            title = h.get_text(strip=True) if h else ""
        if not title:
            title = href.rstrip("/").split("/")[-1].replace("-"," ").title()
        title = re.sub(r"^\d+[\.\)]\s*", "", title).strip()
        if not title or len(title) < 4: continue
        key = title.lower()[:60]
        if key in seen: continue
        cat = _timeout_cat_from_href(href, default_cat)
        is_list = bool(_TIMEOUT_LIST_RE.search(href)) or _is_list_article(title)
        # Only mark as seen if it's an individual event (avoid blocking the key before follow)
        if not is_list: seen.add(key)
        added.append((title, cat, href, is_list))
    return added

def fetch_timeout(location, cfg, tag="", max_list_follows=10):
    """Scrape TimeOut for a city.

    List-article cards (e.g. "50 Best Concerts This Week") are followed to
    extract individual events within them, but only up to *max_list_follows*
    total follow requests to avoid blocking the thread for minutes.
    """
    slug = cfg.get("timeout")
    if not slug: return []
    tprint(f"  [{tag}] → TimeOut ({slug}): {len(TIMEOUT_PAGES)} pages…")
    events, seen = [], set()
    follows_done = 0

    def add(url, default_cat):
        nonlocal follows_done
        r = get(url, min_gap=1.1, silent_404=True)
        if not r: return
        soup = BeautifulSoup(r.text, "lxml")
        for title, cat, href, is_list in _timeout_cards(soup, default_cat, seen):
            if is_list:
                if follows_done >= max_list_follows:
                    continue  # budget exhausted — skip remaining list articles
                follows_done += 1
                # Use a shorter gap (0.7 s) for follow requests — still polite,
                # but the initial listing page already paid the full 1.1 s gap.
                individual = _timeout_follow_list(href, cat, seen, location)
                events.extend(individual)
            else:
                events.append(ev(title, cat, city=location, url=href, source="TimeOut"))

    for path, cat in TIMEOUT_PAGES:
        add(f"https://www.timeout.com/{slug}/{path}", cat)
    if slug == "newyork":
        for path, cat in TIMEOUT_SUBPAGES_NYC:
            add(f"https://www.timeout.com/{slug}/{path}", cat)

    tprint(f"  [{tag}] ✓ TimeOut → {len(events)} (followed {follows_done} list articles)")
    RUNLOG.source("TimeOut", location, len(events))
    return events

# ── 2. Patch.com ──────────────────────────────────────────────────────────────

def _parse_patch(data, location, seen):
    pp, out = data.get("props",{}).get("pageProps",{}), []
    mc    = pp.get("mainContent",{})
    rails = pp.get("rightRail",[])
    def add(item):
        title = (item.get("title") or item.get("shortTitle") or "").strip()
        if not title or title.lower() in seen: return
        seen.add(title.lower())
        body = BeautifulSoup(item.get("body",""),"lxml").get_text(" ")[:200]
        ts   = item.get("displayDateTimestamp") or item.get("created","")
        date_str = parse_date(str(ts)) if str(ts).isdigit() else parse_date(item.get("displayDate",""))
        addr_obj = item.get("address") or {}
        address  = addr_obj.get("display","") if isinstance(addr_obj, dict) else str(addr_obj)
        etype = item.get("eventType","")
        price = "Free" if etype=="free" else ("Paid" if etype=="paid" else "")
        url_  = item.get("canonicalUrl","") or item.get("url","")
        out.append(ev(title, classify(etype+" "+title+" "+body),
            date_str,"",item.get("locationName",""),address,location,price,url_,"Patch.com"))
    for bucket in [mc.get("allEvents",{}), mc.get("promotedEvents",{})]:
        if isinstance(bucket, dict):
            for evlist in bucket.values():
                if isinstance(evlist, list):
                    for item in evlist: add(item)
        elif isinstance(bucket, list):
            for item in bucket: add(item)
    for rail in rails:
        if rail.get("type") == "featuredEvents":
            for item in rail.get("items",[]): add(item)
    return out

def fetch_patch(location, cfg, tag=""):
    slugs = cfg.get("patch",[])
    if not slugs: return []
    tprint(f"  [{tag}] → Patch.com ({len(slugs)} area(s))…")
    seen, all_events = set(), []
    for state, city in slugs:
        r = get(f"https://patch.com/{state}/{city}/calendar", min_gap=1.2)
        if not r: continue
        soup = BeautifulSoup(r.text,"lxml")
        nd = soup.find("script", id="__NEXT_DATA__")
        if not nd: continue
        try:
            evts = _parse_patch(json.loads(nd.string), location, seen)
            all_events.extend(evts)
        except Exception as e:
            tprint(f"  ⚠  Patch parse error ({city}): {e}")
    tprint(f"  [{tag}] ✓ Patch → {len(all_events)}")
    RUNLOG.source("Patch.com", location, len(all_events))
    return all_events

# ── 3. Eventbrite ─────────────────────────────────────────────────────────────

def _eb_next_data(soup, location):
    nd = soup.find("script", id="__NEXT_DATA__")
    if not nd: return []
    events = []
    def walk(obj, depth=0):
        if depth > 10: return
        if isinstance(obj, dict):
            if obj.get("name") and obj.get("start") and "eventbrite.com" in str(obj.get("url","")):
                name = obj["name"].get("text","") if isinstance(obj["name"],dict) else str(obj["name"])
                start = obj["start"].get("local","") if isinstance(obj["start"],dict) else str(obj["start"])
                venue = obj.get("venue",{}) or {}
                vname = venue.get("name","") if isinstance(venue,dict) else ""
                addr  = venue.get("address",{}) or {}
                address = addr.get("address_1","") if isinstance(addr,dict) else ""
                pi    = obj.get("ticket_availability",{}) or {}
                price = "Free" if (isinstance(pi,dict) and pi.get("is_free")) else ""
                events.append(ev(name.strip(), classify(name),
                    parse_date(start), parse_time(start),
                    vname, address, location, price, str(obj.get("url","")), "Eventbrite"))
            for v in obj.values(): walk(v, depth+1)
        elif isinstance(obj, list):
            for i in obj: walk(i, depth+1)
    try: walk(json.loads(nd.string or ""))
    except: pass
    return events

def fetch_eventbrite(location, cfg, max_pages=3, tag=""):
    slug = cfg.get("eb")
    if not slug: return []
    tprint(f"  [{tag}] → Eventbrite ({slug})…")
    all_events, seen = [], set()
    base = f"https://www.eventbrite.com/d/{slug}/events/"
    for page in range(1, max_pages+1):
        r = get(f"{base}?page={page}", min_gap=1.5, hdrs={"Referer":"https://www.eventbrite.com/"})
        if not r: break
        soup = BeautifulSoup(r.text,"lxml")
        batch = _ld_events(soup, location, "Eventbrite") or _eb_next_data(soup, location)
        if not batch:
            for card in soup.select("article,[class*='event-card'],[data-testid*='event']"):
                h    = card.find(["h2","h3","h4","strong"])
                link = card.find("a", href=True)
                if not h or not link: continue
                title = h.get_text(strip=True)
                key   = title.lower()[:60]
                if not title or key in seen or len(key) < 4: continue
                seen.add(key)
                date_el  = card.find("time")
                date_str = parse_date(date_el.get("datetime","")) if date_el else ""
                href = link["href"]
                if not href.startswith("http"): href = "https://www.eventbrite.com" + href
                batch.append(ev(title, classify(title), date_str,"","",
                               "", location,"", href, "Eventbrite"))
        added = 0
        for e in batch:
            key = e.get("Name","").lower()[:60]
            if key and key not in seen:
                seen.add(key); all_events.append(e); added += 1
        if not added: break
    tprint(f"  [{tag}] ✓ Eventbrite → {len(all_events)}")
    RUNLOG.source("Eventbrite", location, len(all_events))
    return all_events

# ── 4. Songkick ───────────────────────────────────────────────────────────────

def fetch_songkick(location, cfg, max_pages=3, tag=""):
    sk_slug = cfg.get("sk_slug")
    if not sk_slug: return []
    tprint(f"  [{tag}] → Songkick ({sk_slug})…")
    all_events, seen = [], set()
    for page in range(1, max_pages+1):
        url = f"https://www.songkick.com/metro-areas/{sk_slug}?page={page}"
        r = get(url, min_gap=1.3, hdrs={
            "Referer": "https://www.songkick.com/",
            "Accept":  "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        })
        if not r: break
        soup = BeautifulSoup(r.text,"lxml")
        batch = _ld_events(soup, location, "Songkick")
        if not batch:
            for art in soup.select("li.event,li[class*='concert'],.concert-event"):
                h    = art.find(["h2","h3","h4","strong"])
                link = art.find("a", href=True)
                if not h: continue
                title = h.get_text(strip=True)
                key   = title.lower()[:60]
                if not title or key in seen or len(key) < 4: continue
                seen.add(key)
                date_el  = art.find("time")
                date_str = parse_date(date_el.get("datetime","") or date_el.get_text()) if date_el else ""
                venue_el = art.find(class_=re.compile(r"venue|location",re.I))
                href = link["href"] if link else ""
                if href and not href.startswith("http"): href = "https://www.songkick.com" + href
                batch.append(ev(title,"Music",date_str,"",
                               venue_el.get_text(strip=True) if venue_el else "",
                               "", location,"", href,"Songkick"))
        added = 0
        for e in batch:
            key = e.get("Name","").lower()[:60]
            if key and key not in seen:
                seen.add(key); all_events.append(e); added += 1
        if not added: break
    tprint(f"  [{tag}] ✓ Songkick → {len(all_events)}")
    RUNLOG.source("Songkick", location, len(all_events))
    return all_events

# ── 5. AllEvents.in ───────────────────────────────────────────────────────────

def fetch_allevents(location, max_pages=2, tag=""):
    city_slug = location.split(",")[0].strip().lower().replace(" ","-")
    tprint(f"  [{tag}] → AllEvents.in ({city_slug})…")
    all_events, seen = [], set()
    for page in range(1, max_pages+1):
        url = (f"https://allevents.in/{city_slug}/all#{page}"
               if page > 1 else f"https://allevents.in/{city_slug}/all")
        r = get(url, min_gap=1.4, hdrs={"Referer":"https://allevents.in/"})
        if not r: break
        soup = BeautifulSoup(r.text,"lxml")
        batch = _ld_events(soup, location, "AllEvents.in")
        if not batch:
            for card in soup.select(".event-item,[class*='event-card'],.item"):
                h    = card.find(["h2","h3","h4","strong"])
                link = card.find("a", href=True)
                if not h: continue
                title = h.get_text(strip=True)
                key   = title.lower()[:60]
                if not title or key in seen or len(key) < 4: continue
                seen.add(key)
                date_el  = card.find("time") or card.find(class_=re.compile(r"date|time",re.I))
                date_str = parse_date(date_el.get_text(strip=True)) if date_el else ""
                href = link["href"] if link else ""
                if href and not href.startswith("http"): href = "https://allevents.in" + href
                batch.append(ev(title, classify(title), date_str,"","","",location,"",href,"AllEvents.in"))
        added = 0
        for e in batch:
            key = e.get("Name","").lower()[:60]
            if key and key not in seen:
                seen.add(key); all_events.append(e); added += 1
        if not added: break
    tprint(f"  [{tag}] ✓ AllEvents → {len(all_events)}")
    RUNLOG.source("AllEvents.in", location, len(all_events))
    return all_events

# ── 6. DoStuff Network ────────────────────────────────────────────────────────

def fetch_dostuff(location, cfg, max_pages=3, tag=""):
    domain = cfg.get("dostuff")
    if not domain: return []
    tprint(f"  [{tag}] → DoStuff ({domain})…")
    all_events, seen = [], set()
    base = f"https://{domain}"

    for page in range(1, max_pages+1):
        # /events/today gives the most reliably populated listing page
        url = f"{base}/events/today" if page == 1 else f"{base}/events/upcoming?page={page}"
        r   = get(url, min_gap=1.2, hdrs={"Referer": base})
        if not r: break
        soup = BeautifulSoup(r.text, "lxml")

        # JSON-LD first
        batch = _ld_events(soup, location, "DoStuff")

        # Card fallback — DoStuff sites share a common card structure
        if not batch:
            selectors = [
                "article.ds-event-card",
                "[class*='event-card']",
                "[class*='EventCard']",
                "article.event",
                ".event-listing",
            ]
            for sel in selectors:
                cards = soup.select(sel)
                if cards:
                    for card in cards:
                        h    = card.find(["h2","h3","h4","strong"])
                        link = card.find("a", href=True)
                        if not h: continue
                        title = h.get_text(strip=True)
                        key   = title.lower()[:60]
                        if not title or key in seen or len(key) < 4: continue
                        seen.add(key)
                        date_el  = card.find("time") or card.find(class_=re.compile(r"date",re.I))
                        date_str = ""
                        if date_el:
                            date_str = parse_date(date_el.get("datetime","") or date_el.get_text(strip=True))
                        venue_el = card.find(class_=re.compile(r"venue|location|place",re.I))
                        href = link["href"] if link else ""
                        if href and not href.startswith("http"): href = base + href
                        batch.append(ev(title, classify(title), date_str,"",
                                       venue_el.get_text(strip=True) if venue_el else "",
                                       "", location,"", href,"DoStuff"))
                    break  # stop trying selectors once one matched

        added = 0
        for e in batch:
            key = e.get("Name","").lower()[:60]
            if key and key not in seen:
                seen.add(key); all_events.append(e); added += 1
        if not added: break

    tprint(f"  [{tag}] ✓ DoStuff → {len(all_events)}")
    RUNLOG.source("DoStuff", location, len(all_events))
    return all_events

# ── 7. Bandsintown ────────────────────────────────────────────────────────────

def fetch_bandsintown(location, cfg, max_pages=2, tag=""):
    city_key = _city_key(location)
    if not city_key or not cfg.get("state"): return []
    slug = _bit_slug(city_key, cfg)
    tprint(f"  [{tag}] → Bandsintown ({slug})…")
    all_events, seen = [], set()
    for page in range(1, max_pages+1):
        url = f"https://www.bandsintown.com/c/{slug}?came_from=257&sort_by_filter=Date&page={page}"
        r   = get(url, min_gap=1.3, hdrs={"Referer":"https://www.bandsintown.com/"})
        if not r: break
        soup = BeautifulSoup(r.text, "lxml")
        batch = _ld_events(soup, location, "Bandsintown")
        if not batch:
            for card in soup.select("[class*='event'],[class*='Event'],[class*='concert']"):
                h    = card.find(["h2","h3","h4","strong"])
                link = card.find("a", href=True)
                if not h: continue
                title = h.get_text(strip=True)
                key   = title.lower()[:60]
                if not title or key in seen or len(key) < 4: continue
                seen.add(key)
                date_el  = card.find("time")
                date_str = parse_date(date_el.get("datetime","")) if date_el else ""
                venue_el = card.find(class_=re.compile(r"venue|location",re.I))
                href = link["href"] if link else ""
                if href and not href.startswith("http"):
                    href = "https://www.bandsintown.com" + href
                batch.append(ev(title,"Music",date_str,"",
                               venue_el.get_text(strip=True) if venue_el else "",
                               "", location,"", href,"Bandsintown"))
        added = 0
        for e in batch:
            key = e.get("Name","").lower()[:60]
            if key and key not in seen:
                seen.add(key); all_events.append(e); added += 1
        if not added: break
    tprint(f"  [{tag}] ✓ Bandsintown → {len(all_events)}")
    RUNLOG.source("Bandsintown", location, len(all_events))
    return all_events

# ── 8. Yelp Events ────────────────────────────────────────────────────────────

def fetch_yelp_events(location, cfg, max_pages=2, tag=""):
    city_key = _city_key(location)
    if not city_key or not cfg.get("state"): return []
    city_name  = location.split(",")[0].strip()
    state_abbr = cfg.get("state","")
    find_loc   = quote_plus(f"{city_name}, {state_abbr}")
    tprint(f"  [{tag}] → Yelp Events ({city_name})…")
    all_events, seen = [], set()
    for page in range(1, max_pages+1):
        offset = (page-1) * 20
        url = (f"https://www.yelp.com/search?cflt=yelpevents&find_loc={find_loc}"
               + (f"&start={offset}" if page > 1 else ""))
        r = get(url, min_gap=1.4, hdrs={"Referer":"https://www.yelp.com/"})
        if not r: break
        soup = BeautifulSoup(r.text, "lxml")
        batch = _ld_events(soup, location, "Yelp Events")
        if not batch:
            for card in soup.select("[class*='event-listing'],[class*='eventCard'],[class*='media-block']"):
                h    = card.find(["h3","h4","strong"])
                link = card.find("a", href=True)
                if not h: continue
                title = h.get_text(strip=True)
                key   = title.lower()[:60]
                if not title or key in seen or len(key) < 4: continue
                seen.add(key)
                date_el  = card.find("time") or card.find(class_=re.compile(r"date",re.I))
                date_str = parse_date(date_el.get_text(strip=True)) if date_el else ""
                venue_el = card.find(class_=re.compile(r"venue|location",re.I))
                href = link["href"] if link else ""
                if href and not href.startswith("http"): href = "https://www.yelp.com" + href
                batch.append(ev(title, classify(title), date_str,"",
                               venue_el.get_text(strip=True) if venue_el else "",
                               "", location,"", href,"Yelp Events"))
        added = 0
        for e in batch:
            key = e.get("Name","").lower()[:60]
            if key and key not in seen:
                seen.add(key); all_events.append(e); added += 1
        if not added: break
    tprint(f"  [{tag}] ✓ Yelp Events → {len(all_events)}")
    RUNLOG.source("Yelp Events", location, len(all_events))
    return all_events

# ── 9. Ticketmaster Discovery API ────────────────────────────────────────────

import os as _os
_TM_KEY = _os.environ.get("TICKETMASTER_API_KEY", "")

def fetch_ticketmaster(location, cfg, max_pages=5, tag="",
                       date_from="", date_to=""):
    """Ticketmaster Discovery API v2 — free, 5 000 calls/day."""
    city_key = _city_key(location)
    if not city_key: return []
    state = cfg.get("state", "")
    city_name = location.split(",")[0].strip()
    tprint(f"  [{tag}] → Ticketmaster ({city_name})…")
    all_events, seen = [], set()
    base = "https://app.ticketmaster.com/discovery/v2/events.json"
    for page in range(0, max_pages):
        params = {
            "apikey":      _TM_KEY,
            "city":        city_name,
            "stateCode":   state,
            "countryCode": "US",
            "size":        50,
            "page":        page,
            "sort":        "date,asc",
        }
        if date_from: params["startDateTime"] = f"{date_from}T00:00:00Z"
        if date_to:   params["endDateTime"]   = f"{date_to}T23:59:59Z"
        try:
            r = SESSION.get(base, params=params, timeout=15)
            r.raise_for_status()
            data = r.json()
        except Exception as e:
            tprint(f"  ⚠  Ticketmaster error: {e}")
            RUNLOG.http_error(base, e)
            break

        items = (data.get("_embedded") or {}).get("events", [])
        if not items:
            break

        for item in items:
            name = item.get("name", "").strip()
            if not name: continue
            key = name.lower()[:60]
            if key in seen: continue
            seen.add(key)

            # Date
            dates  = item.get("dates", {})
            start  = dates.get("start", {})
            date_str = start.get("localDate", "") or parse_date(start.get("dateTime", ""))

            # Venue
            venues   = (item.get("_embedded") or {}).get("venues", [])
            venue_name = venues[0].get("name", "") if venues else ""
            address_parts = []
            if venues:
                v = venues[0]
                line1 = (v.get("address") or {}).get("line1", "")
                if line1: address_parts.append(line1)
                city_v  = (v.get("city") or {}).get("name", "")
                state_v = (v.get("state") or {}).get("stateCode", "")
                if city_v:  address_parts.append(city_v)
                if state_v: address_parts.append(state_v)
            address_str = ", ".join(address_parts)

            # Price
            price_str = ""
            price_ranges = item.get("priceRanges", [])
            if price_ranges:
                lo = price_ranges[0].get("min")
                hi = price_ranges[0].get("max")
                if lo is not None:
                    price_str = f"${lo:.0f}" + (f"–${hi:.0f}" if hi and hi != lo else "")

            # URL
            url = item.get("url", "")

            # Category
            segment = ""
            classifications = item.get("classifications", [])
            if classifications:
                seg_obj = classifications[0].get("segment", {})
                segment = seg_obj.get("name", "")
            cat = {
                "Music": "Music", "Sports": "Sports",
                "Arts & Theatre": "Arts & Theatre",
                "Film": "Film", "Miscellaneous": "Entertainment",
            }.get(segment, classify(name))

            all_events.append(ev(name, cat, date_str, "", venue_name,
                                 address_str, location, price_str, url, "Ticketmaster"))

        page_info = data.get("page", {})
        if page >= page_info.get("totalPages", 1) - 1:
            break

    tprint(f"  [{tag}] ✓ Ticketmaster → {len(all_events)}")
    RUNLOG.source("Ticketmaster", location, len(all_events))
    return all_events


# ── 10. Eventbrite API ───────────────────────────────────────────────────────

_EB_TOKEN = _os.environ.get("EVENTBRITE_TOKEN", "")

def fetch_eventbrite_api(location, cfg, max_pages=5, tag=""):
    """Eventbrite REST API — /v3/events/search/ was deprecated in 2023;
    this function is a no-op until Eventbrite restores public search access."""
    return []

def _fetch_eventbrite_api_disabled(location, cfg, max_pages=5, tag=""):
    """Kept for reference — disabled because /v3/events/search/ returns 404."""
    city_key = _city_key(location)
    if not city_key: return []
    city_name = location.split(",")[0].strip()
    state     = cfg.get("state", "")
    tprint(f"  [{tag}] → Eventbrite API ({city_name})…")
    all_events, seen = [], set()

    # Geocode city to get lat/lon for location-based search
    lat, lon, _ = geocode(location)
    if not lat:
        tprint(f"  [{tag}] ⚠  Eventbrite API: geocode failed, skipping")
        return []

    base = "https://www.eventbriteapi.com/v3/events/search/"
    today_iso = datetime.today().strftime("%Y-%m-%dT00:00:00Z")
    continuation = None

    for _ in range(max_pages):
        params = {
            "location.latitude":  lat,
            "location.longitude": lon,
            "location.within":    "25mi",
            "start_date.range_start": today_iso,
            "expand":   "venue,ticket_availability",
            "sort_by":  "date",
            "page_size": 50,
        }
        if continuation:
            params["continuation"] = continuation

        hdrs = {"Authorization": f"Bearer {_EB_TOKEN}"}
        try:
            r = SESSION.get(base, params=params, headers=hdrs, timeout=20)
            r.raise_for_status()
            data = r.json()
        except Exception as e:
            tprint(f"  ⚠  Eventbrite API error: {e}")
            RUNLOG.http_error(base, e)
            break

        for item in data.get("events", []):
            name = (item.get("name") or {}).get("text", "").strip()
            if not name: continue
            key = name.lower()[:60]
            if key in seen: continue
            seen.add(key)

            date_str = parse_date((item.get("start") or {}).get("local", ""))
            time_str = parse_time((item.get("start") or {}).get("local", ""))

            venue_obj = item.get("venue") or {}
            venue_name = venue_obj.get("name", "")
            addr_obj   = venue_obj.get("address") or {}
            address_str = addr_obj.get("localized_address_display", "")

            # Price
            ticket_av = item.get("ticket_availability") or {}
            price_str = ""
            if ticket_av.get("is_free"):
                price_str = "Free"
            else:
                lo = ticket_av.get("minimum_ticket_price") or {}
                hi = ticket_av.get("maximum_ticket_price") or {}
                lo_v = lo.get("major_value")
                hi_v = hi.get("major_value")
                if lo_v:
                    price_str = f"${lo_v}" + (f"–${hi_v}" if hi_v and hi_v != lo_v else "")

            url = item.get("url", "")
            cat = classify(name)

            all_events.append(ev(name, cat, date_str, time_str, venue_name,
                                 address_str, location, price_str, url, "Eventbrite"))

        pagination = data.get("pagination") or {}
        continuation = pagination.get("continuation")
        if not pagination.get("has_more_items") or not continuation:
            break

    tprint(f"  [{tag}] ✓ Eventbrite API → {len(all_events)}")
    RUNLOG.source("Eventbrite API", location, len(all_events))
    return all_events


# ── 11. National Park Service API ────────────────────────────────────────────

_NPS_KEY = _os.environ.get("NPS_API_KEY", "")

def fetch_nps(location, cfg, tag="", date_from="", date_to=""):
    """NPS Events API — free, covers all national parks by state."""
    state = cfg.get("state", "")
    if not state or not _NPS_KEY: return []
    tprint(f"  [{tag}] → NPS Events ({state})…")
    all_events, seen = [], set()
    start = 0
    while True:
        params = {
            "api_key":   _NPS_KEY,
            "stateCode": state,
            "limit":     50,
            "start":     start,
        }
        if date_from: params["dateStart"] = date_from
        if date_to:   params["dateEnd"]   = date_to
        try:
            r = SESSION.get("https://developer.nps.gov/api/v1/events",
                            params=params, timeout=15)
            r.raise_for_status()
            data = r.json()
        except Exception as e:
            tprint(f"  ⚠  NPS error: {e}")
            RUNLOG.http_error("https://developer.nps.gov/api/v1/events", e)
            break

        items = data.get("data", [])
        if not items:
            break

        for item in items:
            name = item.get("title", "").strip()
            if not name: continue
            key = name.lower()[:60]
            if key in seen: continue
            seen.add(key)

            date_str = parse_date(item.get("date") or item.get("datestart", ""))
            # times is a list like [{"timestart":"09:30 AM",...}]
            times_raw = item.get("times", [])
            if isinstance(times_raw, str):
                try: times_raw = json.loads(times_raw.replace("'", '"'))
                except: times_raw = []
            time_str = times_raw[0].get("timestart", "") if times_raw else ""

            loc      = item.get("location", "")
            park     = item.get("parkfullname") or item.get("organizationname", "")
            url      = item.get("infourl") or item.get("regresurl", "")
            price_str = "Free" if item.get("isfree") == "true" else ""
            cat      = classify(name)

            all_events.append(ev(name, cat, date_str, time_str, park,
                                 loc, location, price_str, url, "NPS"))

        total = int(data.get("total", 0))
        start += len(items)
        if start >= total:
            break

    tprint(f"  [{tag}] ✓ NPS → {len(all_events)}")
    RUNLOG.source("NPS", location, len(all_events))
    return all_events


# ── 12. Socrata Open Data (city permit portals) ───────────────────────────────

# Only Chicago publishes a usable special-events dataset on Socrata.
# Other cities use proprietary portals or don't expose permit data publicly.
_SOCRATA_PORTALS = {
    "chicago": ("data.cityofchicago.org", "xgse-8eg7",
                "event_details", "date", None, "venue_address"),
}

def fetch_socrata_permits(location, cfg, tag=""):
    """Socrata Open Data — special event permits (Chicago only for now)."""
    city_key   = _city_key(location)
    portal_cfg = _SOCRATA_PORTALS.get(city_key)
    if not portal_cfg: return []
    domain, dataset, name_col, date_col, _, loc_col = portal_cfg
    tprint(f"  [{tag}] → City Permits / {domain}…")
    today = datetime.today().strftime("%Y-%m-%d")
    url   = f"https://{domain}/resource/{dataset}.json"
    try:
        r = SESSION.get(url, params={
            "$limit": 200,
            "$where": f"{date_col} >= '{today}'",
            "$order": f"{date_col} ASC",
        }, timeout=20)
        r.raise_for_status()
        rows = r.json()
    except Exception as e:
        tprint(f"  ⚠  Socrata ({domain}) error: {e}")
        RUNLOG.http_error(url, e)
        return []

    all_events, seen = [], set()
    for row in rows:
        name = str(row.get(name_col) or "").strip()
        if not name or len(name) < 4: continue
        key = name.lower()[:60]
        if key in seen: continue
        seen.add(key)
        date_str  = parse_date(row.get(date_col, ""))
        time_str  = row.get("start_time", "")
        venue     = str(row.get("venue") or "").strip()
        loc_str   = str(row.get(loc_col) or "").strip()
        all_events.append(ev(name, classify(name), date_str, time_str,
                             venue, loc_str, location, "", "", "City Permits"))

    tprint(f"  [{tag}] ✓ City Permits → {len(all_events)}")
    RUNLOG.source("City Permits", location, len(all_events))
    return all_events


# ═══════════════════════════════════════════════════════════════════════════════
#  EVENT SOURCES — NYC-SPECIFIC
# ═══════════════════════════════════════════════════════════════════════════════

def fetch_nyc_permitted(limit=2000, tag=""):
    tprint(f"  [{tag}] → NYC Open Data: Permitted Events…")
    today = datetime.today().strftime("%Y-%m-%dT00:00:00.000")
    url = (f"https://data.cityofnewyork.us/resource/tvpp-9vvx.json"
           f"?$limit={limit}&$where=start_date_time>='{today}'&$order=start_date_time+ASC")
    r = get(url, min_gap=1.0, hdrs={"Accept":"application/json"})
    if not r: return []
    skip = {"closure","n/a","none","nan",""}
    out = []
    for item in r.json():
        name = item.get("event_name","").strip()
        if name.lower() in skip: continue
        dr = item.get("start_date_time","")
        out.append(ev(name, classify(item.get("event_type","")+" "+name),
            parse_date(dr), dr[11:16] if len(dr)>10 else "",
            item.get("event_location",""),"",
            item.get("event_borough","New York")+", NY",
            "Varies","","NYC Open Data"))
    tprint(f"  [{tag}] ✓ NYC Permitted → {len(out)}")
    RUNLOG.source("NYC Open Data (Events)", "New York, NY", len(out))
    return out

def fetch_nyc_film_permits(limit=200, tag=""):
    tprint(f"  [{tag}] → NYC Film & Theatre Permits…")
    url = (f"https://data.cityofnewyork.us/resource/tg4x-b46p.json"
           f"?$limit={limit}&$order=enteredon+DESC")
    r = get(url, min_gap=1.0, hdrs={"Accept":"application/json"})
    if not r: return []
    out = []
    for item in r.json():
        etype  = item.get("eventtype","").strip()
        cat    = item.get("category","").strip()
        subcat = item.get("subcategoryname","").strip()
        if not etype: continue
        name = f"{cat}: {subcat}" if subcat and cat else (subcat or etype)
        dr   = item.get("startdatetime","") or item.get("enteredon","")
        out.append(ev(name, classify(cat+" "+subcat+" "+etype),
            parse_date(dr), dr[11:16] if len(dr)>10 else "",
            item.get("parkingheld",""),"",
            item.get("borough","New York")+", NY",
            "Industry/Varies","","NYC Film Permits"))
    tprint(f"  [{tag}] ✓ NYC Film Permits → {len(out)}")
    RUNLOG.source("NYC Film Permits", "New York, NY", len(out))
    return out

def fetch_summerstage(max_pages=6, tag=""):
    tprint(f"  [{tag}] → SummerStage…")
    events, seen = [], set()
    def scrape_page(url):
        r = get(url, min_gap=1.2)
        if not r: return False, None
        soup = BeautifulSoup(r.text,"lxml")
        batch = _ld_events(soup,"New York, NY","SummerStage")
        added = 0
        for e in batch:
            key = e.get("Name","").lower()[:60]
            if key and key not in seen:
                seen.add(key); events.append(e); added += 1
        for art in soup.select("article.type-tribe_events"):
            name_el  = art.find(class_=re.compile(r"event-title",re.I)) or art.find(["h2","h3","h4"])
            link_el  = art.find("a", href=True)
            date_el  = art.find("time")
            name = name_el.get_text(strip=True) if name_el else ""
            if not name or name.lower() in seen: continue
            seen.add(name.lower())
            price_el = art.find(class_=re.compile(r"price|cost",re.I))
            events.append(ev(name, classify(name),
                date_el.get("datetime","") if date_el else "","","","",
                "New York, NY", price_el.get_text(strip=True) if price_el else "",
                link_el["href"] if link_el else "","SummerStage"))
            added += 1
        nxt = soup.find("a", class_=re.compile(r"tribe-events-nav-next|next-events",re.I))
        return added > 0, nxt["href"] if nxt and nxt.get("href") else None
    url = "https://cityparksfoundation.org/events/"
    for _ in range(max_pages):
        had, nxt = scrape_page(url)
        if not had or not nxt: break
        url = nxt
    tprint(f"  [{tag}] ✓ SummerStage → {len(events)}")
    RUNLOG.source("SummerStage", "New York, NY", len(events))
    return events

CAT_MAP_PP = {
    "cat_music":"Music","cat_concert":"Music","cat_art":"Arts & Theatre",
    "cat_dance":"Arts & Theatre","cat_theater":"Arts & Theatre",
    "cat_food":"Food & Drink","cat_wellness":"Sports","cat_sports":"Sports",
    "cat_kids":"Community","cat_community":"Community","cat_education":"Education",
    "cat_film":"Film","cat_nature":"Nature & Parks","cat_free":"Community",
}

def fetch_prospect_park(max_pages=5, tag=""):
    tprint(f"  [{tag}] → Prospect Park…")
    events, seen = [], set()
    base = "https://www.prospectpark.org"
    def scrape_page(url):
        r = get(url, min_gap=1.2)
        if not r: return 0, None
        soup = BeautifulSoup(r.text,"lxml")
        added = 0
        for art in soup.select("article.type-tribe_events"):
            name_el = art.find(class_=re.compile(r"event-title",re.I)) or art.find(["h2","h3","h4"])
            link_el = art.find("a", href=True)
            date_el = art.find("time")
            if not name_el: continue
            name = name_el.get_text(strip=True)
            if not name or name.lower() in seen: continue
            seen.add(name.lower())
            cat = next((CAT_MAP_PP[c] for c in art.get("class",[]) if c in CAT_MAP_PP), classify(name))
            ds  = date_el.get("datetime","") if date_el else ""
            events.append(ev(name,cat,parse_date(ds),parse_time(ds),
                "Prospect Park","Brooklyn","Brooklyn, NY","Free",
                link_el["href"] if link_el else "","Prospect Park"))
            added += 1
        nxt  = soup.find("a",class_=re.compile(r"tribe-events-nav-next|next",re.I),href=True)
        nxt2 = soup.find("a",href=re.compile(r"/events/list/page/\d+/")) if not nxt else None
        nurl = (nxt or nxt2)
        return added, nurl["href"] if nurl else None
    url = f"{base}/events/"
    for _ in range(max_pages):
        added, nxt = scrape_page(url)
        if not added: break
        if not nxt: break
        url = nxt if nxt.startswith("http") else base + nxt
    tprint(f"  [{tag}] ✓ Prospect Park → {len(events)}")
    RUNLOG.source("Prospect Park", "New York, NY", len(events))
    return events

# ═══════════════════════════════════════════════════════════════════════════════
#  ATTRACTION SOURCES
# ═══════════════════════════════════════════════════════════════════════════════

WIKIDATA_CAT = {
    "museum":"Museums","art museum":"Museums","children's museum":"Museums",
    "history museum":"Museums","science museum":"Museums","natural history museum":"Museums",
    "amusement park":"Amusement","theme park":"Amusement","carousel":"Amusement",
    "roller coaster":"Amusement","ferris wheel":"Amusement",
    "observation deck":"Entertainment","tourist attraction":"Entertainment",
    "park":"Nature & Parks","botanical garden":"Nature & Parks",
    "nature reserve":"Nature & Parks","aquarium":"Nature & Parks","zoo":"Nature & Parks",
    "national park":"Nature & Parks","state park":"Nature & Parks",
    "stadium":"Sports","theatre":"Arts & Theatre","theater":"Arts & Theatre",
    "movie theatre":"Film","concert hall":"Music",
    "historic landmark":"Architecture","historic district":"Architecture",
    "lighthouse":"Architecture","historic house":"Architecture","castle":"Architecture",
    "cathedral":"Architecture","church":"Architecture","monument":"Architecture",
}

def fetch_wikidata(lat, lon, radius_km=15, city_label="", tag=""):
    if lat is None: return []
    tprint(f"  [{tag}] → Wikidata (5s pause)…")
    time.sleep(5)
    sparql = f"""
SELECT DISTINCT ?place ?placeLabel ?typeLabel ?website ?article ?coord ?placeDescription WHERE {{
  SERVICE wikibase:around {{
    ?place wdt:P625 ?coord .
    bd:serviceParam wikibase:center "Point({lon} {lat})"^^geo:wktLiteral .
    bd:serviceParam wikibase:radius "{radius_km}" .
  }}
  ?place wdt:P31 ?type .
  VALUES ?type {{
    wd:Q33506 wd:Q207694 wd:Q1007870 wd:Q574915 wd:Q482994
    wd:Q197646 wd:Q1329623 wd:Q22698 wd:Q570116 wd:Q41253
    wd:Q24354 wd:Q14092 wd:Q1107656 wd:Q15078955 wd:Q28564
    wd:Q167346 wd:Q839954 wd:Q44377 wd:Q3947 wd:Q1076486
  }}
  OPTIONAL {{ ?place wdt:P856 ?website }}
  OPTIONAL {{ ?article schema:about ?place ; schema:isPartOf <https://en.wikipedia.org/> }}
  SERVICE wikibase:label {{ bd:serviceParam wikibase:language "en" }}
}}
LIMIT 200
"""
    try:
        resp = SESSION.get("https://query.wikidata.org/sparql",
            params={"query":sparql,"format":"json"},
            headers={"Accept":"application/sparql-results+json","User-Agent":"EventAggregator/8.0"},
            timeout=30)
        resp.raise_for_status()
        bindings = resp.json()["results"]["bindings"]
    except Exception as e:
        tprint(f"  ⚠  Wikidata: {e}"); return []
    places, seen = [], set()
    for b in bindings:
        name  = b.get("placeLabel",{}).get("value","").strip()
        ptype = b.get("typeLabel",{}).get("value","").strip()
        web   = b.get("website",{}).get("value","")
        art   = b.get("article",{}).get("value","")
        desc  = b.get("placeDescription",{}).get("value","")
        if not name or name in seen or re.match(r"^Q\d+$", name): continue
        # Skip entries with no online presence — they're typically generic/unremarkable
        if not (web or art): continue
        seen.add(name)
        # Parse WKT "Point(lon lat)" — WKT is lon-first, opposite of lat/lon convention
        wkt = b.get("coord",{}).get("value","")
        wm  = re.match(r"Point\(([+-]?\d+\.?\d*)\s+([+-]?\d+\.?\d*)\)", wkt or "")
        wlat, wlon = (float(wm.group(2)), float(wm.group(1))) if wm else (None, None)
        cat = WIKIDATA_CAT.get(ptype.lower(), classify(name+" "+ptype))
        places.append(att(name, cat, ptype or "Attraction", "", city_label,
                          web or art, "Wikidata", lat=wlat, lon=wlon, desc=desc))
    tprint(f"  [{tag}] ✓ Wikidata → {len(places)}")
    RUNLOG.source("Wikidata", city_label, len(places))
    return places

OSM_MAP = {
    "attraction":"Entertainment","museum":"Museums","theme_park":"Amusement",
    "viewpoint":"Entertainment","zoo":"Nature & Parks","aquarium":"Nature & Parks",
    "gallery":"Arts & Theatre","park":"Nature & Parks","nature_reserve":"Nature & Parks",
    "garden":"Nature & Parks","stadium":"Sports","theatre":"Arts & Theatre",
    "cinema":"Film","arts_centre":"Arts & Theatre","monument":"Architecture",
    "memorial":"Architecture","lighthouse":"Architecture","historic":"Architecture",
    "artwork":"Arts & Theatre",
}

def fetch_osm(lat, lon, radius_m=8000, limit=150, city_label="", tag=""):
    if lat is None: return []
    tprint(f"  [{tag}] → OpenStreetMap Overpass…")
    query = f"""[out:json][timeout:25];
(
  node["tourism"~"attraction|museum|theme_park|zoo|aquarium|gallery|viewpoint|artwork"](around:{radius_m},{lat},{lon});
  node["amenity"~"theatre|cinema|arts_centre"](around:{radius_m},{lat},{lon});
  node["leisure"~"park|nature_reserve|garden|stadium"](around:{radius_m},{lat},{lon});
  node["historic"~"monument|memorial|landmark|castle|lighthouse"](around:{radius_m},{lat},{lon});
);
out {limit};"""
    r = post_req("https://overpass-api.de/api/interpreter",{"data":query},
                 hdrs={"User-Agent":"EventAggregator/8.0"})
    if not r: return []
    places, seen = [], set()
    for el in r.json().get("elements",[]):
        tags = el.get("tags",{})
        name = tags.get("name","").strip()
        if not name or name in seen: continue
        t=tags.get("tourism",""); l=tags.get("leisure","")
        a=tags.get("amenity",""); h=tags.get("historic","")
        web  = tags.get("website", tags.get("contact:website",""))
        wiki = tags.get("wikipedia","")
        addr = " ".join(filter(None,[tags.get("addr:housenumber",""),tags.get("addr:street","")]))
        # Skip generic "attraction" nodes with no website, wikipedia, or address
        # — they're usually unmapped placeholder nodes, not real landmarks
        if t == "attraction" and not (web or wiki or addr): continue
        seen.add(name)
        cat  = OSM_MAP.get(t) or OSM_MAP.get(l) or OSM_MAP.get(a) or OSM_MAP.get(h) or classify(name)
        elat = el.get("lat"); elon = el.get("lon")
        places.append(att(name, cat, t or l or a or h or "Attraction", addr,
                          tags.get("addr:city","") or city_label, web,
                          "OpenStreetMap", lat=elat, lon=elon))
    tprint(f"  [{tag}] ✓ OpenStreetMap → {len(places)}")
    RUNLOG.source("OpenStreetMap", city_label, len(places))
    return places

def fetch_nyc_cultural_orgs(limit=300, tag=""):
    tprint(f"  [{tag}] → NYC Cultural Organizations…")
    r = get(f"https://data.cityofnewyork.us/resource/u35m-9t32.json?$limit={limit}",
            min_gap=1.0, hdrs={"Accept":"application/json"})
    if not r: return []
    places, seen = [], set()
    for item in r.json():
        name = item.get("organization_name","").strip()
        if not name or name in seen: continue
        seen.add(name)
        disc = item.get("discipline","")
        cat  = classify(disc+" "+name)
        if "theater" in disc.lower() or "theatre" in disc.lower(): cat="Arts & Theatre"
        elif "music" in disc.lower():  cat="Music"
        elif "museum" in disc.lower(): cat="Museums"
        elif "dance" in disc.lower():  cat="Arts & Theatre"
        elif "film" in disc.lower():   cat="Film"
        places.append(att(name,cat,disc or "Cultural Org",
                          item.get("address",""),
                          f"{item.get('city','New York')}, {item.get('state','NY')}",
                          "","NYC Cultural Orgs"))
    tprint(f"  [{tag}] ✓ NYC Cultural Orgs → {len(places)}")
    RUNLOG.source("NYC Cultural Orgs", "New York, NY", len(places))
    return places

def fetch_nyc_farmers_markets(limit=200, tag=""):
    tprint(f"  [{tag}] → NYC Farmers Markets…")
    r = get(f"https://data.cityofnewyork.us/resource/8vwk-6iz2.json?$limit={limit}",
            min_gap=1.0, hdrs={"Accept":"application/json"})
    if not r: return []
    places = []
    for item in r.json():
        name = item.get("marketname","").strip()
        if not name: continue
        days  = item.get("daysoperation","")
        hours = item.get("hoursoperations","")
        yr    = "Year-round" if item.get("open_year_round","").lower()=="yes" else "Seasonal"
        ebt   = " | EBT" if item.get("accepts_ebt","").lower()=="yes" else ""
        label = f"{name} ({days}, {hours})" if days else name
        places.append(att(label,"Food & Drink",f"Farmers Market — {yr}{ebt}",
                          item.get("streetaddress",""),
                          f"{item.get('borough','New York')}, NY","","NYC Farmers Markets"))
    tprint(f"  [{tag}] ✓ NYC Farmers Markets → {len(places)}")
    RUNLOG.source("NYC Farmers Markets", "New York, NY", len(places))
    return places

# ═══════════════════════════════════════════════════════════════════════════════
#  PARALLEL CITY FETCHER
# ═══════════════════════════════════════════════════════════════════════════════

def fetch_city(location, skip_attractions=False, source_workers=6,
               date_from="", date_to=""):
    """
    Fetch all events and attractions for one city.
    Sources run concurrently (different domains → no rate conflicts).
    date_from / date_to: ISO date strings (YYYY-MM-DD) for optional range filter.
    Returns (events, attractions).
    """
    city_key = _city_key(location)
    cfg      = CITY_CONFIG.get(city_key, {})
    is_nyc   = city_key == "new york"
    tag      = location.split(",")[0]  # short label for log lines

    tprint(f"\n{'─'*60}")
    tprint(f"  📍  {location}")
    if date_from or date_to:
        tprint(f"  📅  {date_from or 'any'} → {date_to or 'any'}")
    tprint(f"{'─'*60}")

    # ── Build task list ───────────────────────────────────────────────────────
    event_tasks = []
    if is_nyc:
        event_tasks += [
            partial(fetch_nyc_permitted,  2000,    tag=tag),
            partial(fetch_nyc_film_permits, 200,   tag=tag),
            partial(fetch_summerstage,    6,        tag=tag),
            partial(fetch_prospect_park,  5,        tag=tag),
        ]
    if cfg:
        event_tasks += [
            partial(fetch_timeout,        location, cfg, tag=tag),
            partial(fetch_patch,          location, cfg, tag=tag),
            partial(fetch_eventbrite,     location, cfg, tag=tag),
            partial(fetch_eventbrite_api, location, cfg, tag=tag),
            partial(fetch_songkick,       location, cfg, tag=tag),
            partial(fetch_dostuff,        location, cfg, tag=tag),
            partial(fetch_yelp_events,    location, cfg, tag=tag),
            partial(fetch_ticketmaster,   location, cfg,
                    date_from=date_from, date_to=date_to, tag=tag),
            partial(fetch_nps,            location, cfg,
                    date_from=date_from, date_to=date_to, tag=tag),
            partial(fetch_socrata_permits,location, cfg, tag=tag),
            # Bandsintown disabled — returns 403 (actively blocks scrapers)
        ]
    event_tasks.append(partial(fetch_allevents, location, tag=tag))

    # ── Run event sources in parallel ─────────────────────────────────────────
    all_events = []
    with ThreadPoolExecutor(max_workers=source_workers) as pool:
        futures = {pool.submit(fn): getattr(fn, "func", fn).__name__
                   for fn in event_tasks}
        for future in as_completed(futures):
            fname = futures[future]
            try:
                result = future.result()
                if result: all_events.extend(result)
            except Exception as e:
                tprint(f"  ⚠  Source error ({fname}): {e}")
                RUNLOG.source(fname, location, 0, error=str(e))

    # ── Attractions (geocode first, then parallel Wikidata+OSM) ───────────────
    all_attractions = []
    if not skip_attractions:
        lat, lon, display = geocode(location)
        if lat:
            tprint(f"  [{tag}] 📍 {display[:65]}")
            att_tasks = [
                partial(fetch_wikidata, lat, lon, 15, location, tag=tag),
                partial(fetch_osm,      lat, lon, 8000, 150, location, tag=tag),
            ]
            if is_nyc:
                att_tasks += [
                    partial(fetch_nyc_cultural_orgs,  300, tag=tag),
                    partial(fetch_nyc_farmers_markets, 200, tag=tag),
                ]
            with ThreadPoolExecutor(max_workers=4) as pool:
                att_futures = {pool.submit(fn): getattr(fn, "func", fn).__name__
                               for fn in att_tasks}
                for future in as_completed(att_futures):
                    fname = att_futures[future]
                    try:
                        result = future.result()
                        if result: all_attractions.extend(result)
                    except Exception as e:
                        tprint(f"  ⚠  Attraction source error ({fname}): {e}")
                        RUNLOG.source(fname, location, 0, error=str(e))
        else:
            tprint(f"  [{tag}] ⚠  Geocoding failed — attraction sources skipped")

    return all_events, all_attractions

# ═══════════════════════════════════════════════════════════════════════════════
#  DEDUP
# ═══════════════════════════════════════════════════════════════════════════════

def dedup(lst, filter_articles=False, validate_events=False,
          date_from="", date_to=""):
    seen, out = set(), []
    for x in lst:
        name = (x.get("Name") or "").strip()
        k = name.lower()[:60]
        if not k or k in seen: continue
        if filter_articles and _is_list_article(name): continue
        if validate_events and not _is_valid_event(x): continue
        d = x.get("Date") or ""
        if date_from and d and d < date_from: continue
        if date_to   and d and d > date_to:   continue
        seen.add(k); out.append(x)
    return out

# ═══════════════════════════════════════════════════════════════════════════════
#  EXCEL OUTPUT
# ═══════════════════════════════════════════════════════════════════════════════

ECOLS = ["#","Name","Category","Date","Time","Venue","Address","City","Price","URL","Source"]
ACOLS = ["#","Name","Category","Type","Description","Address","City","URL","Source"]

def _title(ws, text, ncols, bg, sz=13):
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    hdr(ws["A1"], bg=bg, sz=sz)
    ws["A1"].value = text
    ws.row_dimensions[1].height = 28

def build_summary(wb, events, attractions, label):
    ws = wb.create_sheet("📊 Summary", 0)
    ws.sheet_properties.tabColor = "833C00"
    _title(ws, "Event & Activities Aggregator — Summary", 3, "833C00", 14)

    def kv(row, k, v):
        c1 = ws.cell(row=row, column=1, value=k)
        c2 = ws.cell(row=row, column=2, value=str(v))
        c1.font = Font(bold=True, name="Arial", size=10)
        c2.font = Font(name="Arial", size=10)
        for c in [c1, c2]:
            c.fill   = PatternFill("solid", fgColor="F2F2F2")
            c.border = _b()

    cities = sorted({e.get("City","") for e in events+attractions if e.get("City","")})
    kv(2, "Search",                       label)
    kv(3, "Generated On",                 datetime.now().strftime("%Y-%m-%d %H:%M"))
    kv(4, "Cities Covered",               len(cities))
    kv(5, "Total Events",                 len(events))
    kv(6, "Total Attractions / Things To Do", len(attractions))
    kv(7, "Grand Total",                  len(events)+len(attractions))

    row = 9
    for htext, bg_, items, field in [
        ("Events by City",          "1F5C8B", events,               "City"),
        ("Events by Category",      "2E75B6", events,               "Category"),
        ("Attractions by Category", "375623", attractions,          "Category"),
        ("Data Sources",            "595959", events+attractions,   "Source"),
    ]:
        ws.merge_cells(f"A{row}:C{row}")
        hdr(ws.cell(row, 1, htext), bg=bg_)
        ws.row_dimensions[row].height = 20
        row += 1
        for val, cnt in Counter(x[field] for x in items).most_common():
            bg = CAT_COLOURS.get(val,"F5F5F5") if field=="Category" else "F2F2F2"
            c1 = ws.cell(row, 1, val); c2 = ws.cell(row, 2, cnt)
            dat(c1, bg); dat(c2, bg)
            c2.alignment = Alignment(horizontal="center", vertical="center")
            row += 1
        row += 1

    set_widths(ws, {"A":32,"B":14,"C":14})

def _link_cell(cell, url):
    """Style a cell as a short hyperlink: show domain as display text."""
    if not url: return
    try:
        domain = urlparse(url).netloc.replace("www.","")
    except Exception:
        domain = "Link"
    cell.hyperlink = url
    cell.value     = domain or "Link"
    cell.font      = Font(color="0563C1", underline="single", name="Arial", size=10)

def _row_height(*cell_pairs, min_h=18, max_h=90):
    """Estimate row height in pts so all wrapped cells stay visible.
    Pass alternating (text, col_width_chars) pairs for each wrapped column."""
    max_lines = 1
    it = iter(cell_pairs)
    for text, width in zip(it, it):
        if not text: continue
        lines = math.ceil(len(str(text)) / max(1, int(width)))
        max_lines = max(max_lines, lines)
    return min(max(min_h, max_lines * 14), max_h)

def build_events_sheet(wb, events, label):
    ws = wb.create_sheet("📅 Events")
    ws.sheet_properties.tabColor = "2E75B6"
    _title(ws, f"Events — {label}", len(ECOLS), "1F3864")
    for c, n in enumerate(ECOLS, 1): hdr(ws.cell(2, c, n))
    ws.row_dimensions[2].height = 22
    for i, e in enumerate(events, 1):
        row = i + 2
        bg  = CAT_COLOURS.get(e.get("Category","Other"), "F5F5F5")
        url = e.get("URL","")
        vals = [i, e.get("Name"), e.get("Category"), e.get("Date"), e.get("Time"),
                e.get("Venue"), e.get("Address"), e.get("City"),
                e.get("Price"), url, e.get("Source")]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row, c, v); dat(cell, bg)
            if c == 1:  cell.alignment = Alignment(horizontal="center", vertical="center")
            if c == 10: _link_cell(cell, url)
        # Consider Name (B=38) and Venue (F=26) — whichever wraps more sets the height
        ws.row_dimensions[row].height = _row_height(
            e.get("Name",""), 38, e.get("Venue",""), 26)
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(ECOLS))}2"
    set_widths(ws,{"A":5,"B":38,"C":16,"D":12,"E":7,
                   "F":26,"G":22,"H":18,"I":10,"J":18,"K":18})

def build_attractions_sheet(wb, attractions, label):
    ws = wb.create_sheet("📍 Things To Do")
    ws.sheet_properties.tabColor = "375623"
    _title(ws, f"Attractions & Things To Do — {label}", len(ACOLS), "375623")
    for c, n in enumerate(ACOLS, 1): hdr(ws.cell(2, c, n), bg="375623")
    ws.row_dimensions[2].height = 22
    url_col = ACOLS.index("URL") + 1  # 1-based column index for URL
    for i, pl in enumerate(attractions, 1):
        row = i + 2
        bg  = CAT_COLOURS.get(pl.get("Category","Other"),"F5F5F5")
        url = pl.get("URL","")
        vals = [i, pl.get("Name"), pl.get("Category"), pl.get("Type"),
                pl.get("Description",""), pl.get("Address"), pl.get("City"),
                url, pl.get("Source")]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row, c, v); dat(cell, bg)
            if c == 1: cell.alignment = Alignment(horizontal="center", vertical="center")
            if c == url_col: _link_cell(cell, url)
        ws.row_dimensions[row].height = _row_height(
            pl.get("Name",""), 38, pl.get("Description",""), 40)
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(ACOLS))}2"
    set_widths(ws,{"A":5,"B":38,"C":16,"D":22,"E":40,"F":22,"G":18,"H":18,"I":16})

# ═══════════════════════════════════════════════════════════════════════════════
#  MAP OUTPUT
# ═══════════════════════════════════════════════════════════════════════════════

# Category → pin color (hex) for the Leaflet map
_MAP_COLORS = {
    "Music":           "#4A90D9",
    "Sports":          "#5CB85C",
    "Arts & Theatre":  "#9B59B6",
    "Food & Drink":    "#E67E22",
    "Community":       "#F1C40F",
    "Nature & Parks":  "#27AE60",
    "Museums":         "#8E44AD",
    "Amusement":       "#E74C3C",
    "Shopping":        "#F39C12",
    "Entertainment":   "#E91E63",
    "Architecture":    "#95A5A6",
    "Film":            "#3498DB",
    "Education":       "#16A085",
    "Other":           "#7F8C8D",
}

def build_map(events, attractions, label, out_path):
    """Generate a self-contained Leaflet + MarkerCluster HTML map.

    Architecture:
      - ONE event cluster group  (all events)
      - ONE attraction cluster group (all attractions, real GPS coords)
    Category toggles add/remove individual marker objects from those two
    cluster groups rather than having 15 separate per-category clusters
    all piling on top of each other at the same city-center point.
    """
    geo = dict(_GEO_CACHE)

    def city_center(city_str):
        for k, (la, lo, _) in geo.items():
            if city_str.lower() in k.lower() or k.lower() in city_str.lower():
                return la, lo
        return None, None

    ev_records = []
    for e in events:
        clat, clon = city_center(e.get("City",""))
        if clat is None: continue
        ev_records.append({
            "n":     e.get("Name",""),
            "cat":   e.get("Category","Other"),
            "date":  e.get("Date",""),
            "time":  e.get("Time",""),
            "venue": e.get("Venue",""),
            "city":  e.get("City",""),
            "price": e.get("Price",""),
            "url":   e.get("URL",""),
            "src":   e.get("Source",""),
            "lat": clat, "lon": clon,
        })

    att_records = []
    for a in attractions:
        alat = a.get("lat"); alon = a.get("lon")
        if alat is None:
            alat, alon = city_center(a.get("City",""))
        if alat is None: continue
        att_records.append({
            "n":    a.get("Name",""),
            "cat":  a.get("Category","Other"),
            "type": a.get("Type",""),
            "desc": a.get("Description",""),
            "addr": a.get("Address",""),
            "city": a.get("City",""),
            "url":  a.get("URL",""),
            "lat": alat, "lon": alon,
        })

    colors_js  = json.dumps(_MAP_COLORS)
    events_js  = json.dumps(ev_records,  ensure_ascii=False)
    attribs_js = json.dumps(att_records, ensure_ascii=False)
    gen_ts     = datetime.now().strftime("%Y-%m-%d %H:%M")

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Event Map — {label}</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
<link rel="stylesheet" href="https://unpkg.com/leaflet.markercluster@1.5.3/dist/MarkerCluster.css"/>
<link rel="stylesheet" href="https://unpkg.com/leaflet.markercluster@1.5.3/dist/MarkerCluster.Default.css"/>
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<script src="https://unpkg.com/leaflet.markercluster@1.5.3/dist/leaflet.markercluster.js"></script>
<style>
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{ font-family: Arial, sans-serif; display: flex; flex-direction: column; height: 100vh; }}
#header {{
  background: #1F3864; color: #fff; padding: 8px 16px;
  display: flex; align-items: center; justify-content: space-between; flex-shrink: 0;
}}
#header h1 {{ font-size: 15px; font-weight: bold; }}
#header small {{ font-size: 11px; opacity: 0.75; }}
#controls {{
  background: #f5f5f5; border-bottom: 1px solid #ddd;
  padding: 6px 12px; display: flex; flex-wrap: wrap; gap: 5px; flex-shrink: 0;
}}
.layer-btn {{
  padding: 3px 10px; border-radius: 12px; border: none;
  font-size: 11px; cursor: pointer; font-weight: bold; color: #fff;
  transition: opacity 0.15s; white-space: nowrap;
}}
.layer-btn.off {{ opacity: 0.28; }}
#map {{ flex: 1; }}
/* Cluster bubble styles */
.ev-cluster, .att-cluster {{
  border-radius: 50%; display: flex; align-items: center; justify-content: center;
  font-weight: bold; border: 3px solid #fff;
  box-shadow: 0 2px 8px rgba(0,0,0,.5); color: #fff;
}}
.ev-cluster  {{ background: #1F3864; }}
.att-cluster {{ background: #375623; }}
/* Popup */
.popup-box {{ min-width: 200px; max-width: 270px; font-family: Arial,sans-serif; line-height: 1.4; }}
.popup-title {{ font-weight: bold; font-size: 13px; margin-bottom: 5px; }}
.popup-row {{ font-size: 11px; color: #444; margin: 2px 0; }}
.popup-src  {{ font-size: 10px; color: #aaa; margin-top: 4px; }}
.popup-link a {{ font-size: 11px; color: #0563C1; font-weight: bold; text-decoration: none; }}
.popup-link a:hover {{ text-decoration: underline; }}
</style>
</head>
<body>
<div id="header">
  <h1>🗺 Event Map — {label}</h1>
  <small>Generated {gen_ts} &nbsp;·&nbsp; {len(ev_records)} events &nbsp;·&nbsp; {len(att_records)} attractions</small>
</div>
<div id="controls"></div>
<div id="map"></div>
<script>
const COLORS  = {colors_js};
const EVENTS  = {events_js};
const ATTRIBS = {attribs_js};

const map = L.map('map');
L.tileLayer('https://{{s}}.basemaps.cartocdn.com/rastertiles/voyager/{{z}}/{{x}}/{{y}}{{r}}.png', {{
  attribution: '© <a href="https://openstreetmap.org">OpenStreetMap</a> © <a href="https://carto.com">CARTO</a>',
  subdomains: 'abcd', maxZoom: 19
}}).addTo(map);

function esc(s) {{
  return (s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
}}

// ── Two cluster groups — one for events, one for attractions ────────────────
//    This is the key fix: not 15 separate category clusters that all pile up
//    at city center. Category toggling works by adding/removing individual
//    marker objects from these two groups.

function makeCluster(cls) {{
  return L.markerClusterGroup({{
    maxClusterRadius: 70,
    spiderfyOnMaxZoom: true,
    spiderfyDistanceMultiplier: 2,
    showCoverageOnHover: false,
    chunkedLoading: true,
    iconCreateFunction: function(cluster) {{
      const n  = cluster.getChildCount();
      const sz = n < 20 ? 36 : n < 100 ? 44 : 54;
      return L.divIcon({{
        html: '<div class="' + cls + '" style="width:' + sz + 'px;height:' + sz + 'px;font-size:' + Math.round(sz*0.33) + 'px">' + n + '</div>',
        className: '', iconSize: [sz,sz], iconAnchor: [sz/2,sz/2]
      }});
    }}
  }});
}}

const evCluster  = makeCluster('ev-cluster').addTo(map);
const attCluster = makeCluster('att-cluster').addTo(map);

// ── Track markers per category so toggles can add/remove them ──────────────
const evByCat  = {{}};   // cat  -> [marker, ...]
const attByCat = {{}};   // cat  -> [marker, ...]
const catEvOn  = {{}};   // cat  -> bool (events)
const catAttOn = {{}};   // cat  -> bool (attractions)

// ── Plot events ─────────────────────────────────────────────────────────────
EVENTS.forEach(function(e) {{
  const color = COLORS[e.cat] || '#888';
  const m = L.circleMarker([e.lat, e.lon], {{
    radius: 9, fillColor: color, color: '#fff', weight: 2,
    opacity: 1, fillOpacity: 1
  }});
  let p = '<div class="popup-box">'
        + '<div class="popup-title" style="color:'+color+'">' + esc(e.n) + '</div>';
  if (e.date||e.time) p += '<div class="popup-row">📅 ' + esc([e.date,e.time].filter(Boolean).join(' · ')) + '</div>';
  if (e.venue)  p += '<div class="popup-row">🏛 ' + esc(e.venue) + '</div>';
  if (e.city)   p += '<div class="popup-row">📍 ' + esc(e.city)  + '</div>';
  if (e.price)  p += '<div class="popup-row">💰 ' + esc(e.price) + '</div>';
  p += '<div class="popup-src">Source: ' + esc(e.src) + '</div>';
  if (e.url) p += '<div class="popup-link" style="margin-top:6px"><a href="' + esc(e.url) + '" target="_blank">Details →</a></div>';
  p += '</div>';
  m.bindPopup(p);
  evCluster.addLayer(m);
  (evByCat[e.cat] = evByCat[e.cat] || []).push(m);
  catEvOn[e.cat] = true;
}});

// ── Plot attractions ─────────────────────────────────────────────────────────
ATTRIBS.forEach(function(a) {{
  const color = COLORS[a.cat] || '#888';
  const m = L.circleMarker([a.lat, a.lon], {{
    radius: 7, fillColor: color, color: '#333', weight: 1.5,
    opacity: 1, fillOpacity: 1
  }});
  let p = '<div class="popup-box">'
        + '<div class="popup-title" style="color:'+color+'">' + esc(a.n) + '</div>';
  if (a.type) p += '<div class="popup-row">🏷 ' + esc(a.cat) + ' · ' + esc(a.type) + '</div>';
  if (a.desc) p += '<div class="popup-row" style="font-style:italic;color:#555">' + esc(a.desc) + '</div>';
  if (a.addr) p += '<div class="popup-row">📍 ' + esc(a.addr) + '</div>';
  if (a.city) p += '<div class="popup-row">🏙 ' + esc(a.city) + '</div>';
  if (a.url)  p += '<div class="popup-link" style="margin-top:6px"><a href="' + esc(a.url) + '" target="_blank">Learn more →</a></div>';
  p += '</div>';
  m.bindPopup(p);
  attCluster.addLayer(m);
  (attByCat[a.cat] = attByCat[a.cat] || []).push(m);
  catAttOn[a.cat] = true;
}});

// ── Fit bounds to attractions (spread across city) + city centers ───────────
const fitCoords = [
  ...ATTRIBS.map(a=>[a.lat,a.lon]),
  ...EVENTS.map(e=>[e.lat,e.lon])
].filter(c=>c[0]!=null);
if (fitCoords.length) map.fitBounds(fitCoords, {{padding:[30,30]}});
else map.setView([39.5,-98.35],4);

// ── Category toggle buttons ───────────────────────────────────────────────────
const ctrl = document.getElementById('controls');

// Helper: add a labelled divider span
function divider(text) {{
  const s = document.createElement('span');
  s.style.cssText = 'font-size:10px;color:#888;align-self:center;padding:0 4px;white-space:nowrap';
  s.textContent = text;
  ctrl.appendChild(s);
}}

function addBtn(label, color, onToggle, startOn) {{
  const btn = document.createElement('button');
  btn.className = 'layer-btn';
  btn.style.background = color;
  btn.textContent = label;
  let on = startOn;
  btn.onclick = function() {{
    on = !on;
    onToggle(on);
    btn.classList.toggle('off', !on);
  }};
  ctrl.appendChild(btn);
}}

// "All Events" master toggle
divider('Events:');
addBtn('All Events', '#1F3864', function(on) {{
  on ? map.addLayer(evCluster) : map.removeLayer(evCluster);
}}, true);

// Per-category event toggles
const evCats = Object.keys(evByCat).sort();
evCats.forEach(function(cat) {{
  const color = COLORS[cat] || '#888';
  addBtn('📅 '+cat, color, function(on) {{
    const ms = evByCat[cat] || [];
    ms.forEach(m => on ? evCluster.addLayer(m) : evCluster.removeLayer(m));
  }}, true);
}});

// "All Attractions" master toggle + per-category
divider('  Attractions:');
addBtn('All Attractions', '#375623', function(on) {{
  on ? map.addLayer(attCluster) : map.removeLayer(attCluster);
}}, true);

const attCats = Object.keys(attByCat).sort();
attCats.forEach(function(cat) {{
  const color = COLORS[cat] || '#888';
  addBtn('📍 '+cat, color, function(on) {{
    const ms = attByCat[cat] || [];
    ms.forEach(m => on ? attCluster.addLayer(m) : attCluster.removeLayer(m));
  }}, true);
}});
</script>
</body>
</html>"""

    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)

# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    print("="*68)
    print("  Event & Activities Aggregator v8  |  US National Coverage")
    print("="*68)
    print()
    print("  MODES:")
    print("  [1] Single city        e.g.  New York, NY")
    print("  [2] Multi-city         e.g.  Boston, MA; Chicago, IL; Miami, FL")
    print("  [3] Region preset")
    print()

    mode = input("Mode [1/2/3]: ").strip()

    cities = []
    label  = ""

    if mode == "3":
        print()
        for r_name, r_cities in REGIONS.items():
            preview = ", ".join(c.title() for c in r_cities[:5])
            print(f"  {r_name:<12} — {preview}{'…' if len(r_cities)>5 else ''}")
        region_in = input("\nRegion name: ").strip().lower()
        if region_in not in REGIONS:
            print(f"Unknown region '{region_in}'. Options: {', '.join(REGIONS.keys())}")
            sys.exit(1)
        cities = []
        for ck in REGIONS[region_in]:
            cfg   = CITY_CONFIG.get(ck, {})
            state = cfg.get("state","")
            cities.append(f"{ck.title()}, {state}" if state else ck.title())
        label = f"{region_in.title()} Region ({len(cities)} cities)"

    elif mode == "2":
        raw    = input("Cities (semicolons, e.g. 'Boston, MA; Chicago, IL'): ").strip()
        cities = [c.strip() for c in raw.split(";") if c.strip()]
        label  = f"Multi-city: {', '.join(cities)}"

    else:
        city_in = input("City (e.g. 'New York, NY'): ").strip() or "New York, NY"
        cities  = [city_in]
        label   = city_in

    skip_attr_in = input("Skip attractions? (faster for multi-city runs) [y/N]: ").strip().lower()
    skip_attractions = skip_attr_in == "y"

    city_workers = min(4, len(cities))   # concurrent cities
    src_workers  = 6                      # concurrent sources per city

    print(f"\n{'─'*68}")
    print(f"  📡  {len(cities)} city/cities | city workers: {city_workers} | source workers: {src_workers}")
    print(f"  Sources: TimeOut · Patch · Eventbrite · Songkick · DoStuff")
    print(f"           AllEvents · Bandsintown · Yelp Events")
    if not skip_attractions:
        print(f"  Attractions: Wikidata · OpenStreetMap")
    print(f"{'─'*68}\n")

    t0 = time.time()
    all_events, all_attractions = [], []
    city_lock = threading.Lock()

    def run_city(city_loc):
        evts, atts = fetch_city(city_loc, skip_attractions=skip_attractions,
                                source_workers=src_workers)
        with city_lock:
            all_events.extend(evts)
            all_attractions.extend(atts)
        return city_loc, len(evts), len(atts)

    with ThreadPoolExecutor(max_workers=city_workers) as pool:
        futures = {pool.submit(run_city, city): city for city in cities}
        for future in as_completed(futures):
            try:
                city_loc, ne, na = future.result()
                tprint(f"\n  ✅  {city_loc}: {ne} events, {na} attractions")
            except Exception as e:
                tprint(f"\n  ⚠  {futures[future]} failed: {e}")

    raw_event_count  = len(all_events)
    all_events       = dedup(all_events, filter_articles=True, validate_events=True)
    all_attractions  = dedup(all_attractions)
    dropped          = raw_event_count - len(all_events)
    tprint(f"\n  🔍  Validation: {raw_event_count} raw → {len(all_events)} kept"
           f"  ({dropped} dropped: past/undated/no-url/songs/articles)")
    all_events.sort(key=lambda e: (e.get("Date","") or "9999", e.get("City","")))

    elapsed = time.time() - t0
    print(f"\n{'─'*68}")
    print(f"  ✅  {len(all_events)} events  |  {len(all_attractions)} attractions")
    print(f"      {len(cities)} city/cities  |  {elapsed:.0f}s elapsed")
    print(f"{'─'*68}")

    if not all_events and not all_attractions:
        print("\n✗  No data found."); sys.exit(1)

    # ── Write run report ──────────────────────────────────────────────────────
    report_path = f"run_report_{datetime.today().strftime('%Y%m%d_%H%M%S')}.txt"
    RUNLOG.write(report_path, label, elapsed)

    safe = re.sub(r"[^\w]","_", label)[:40]
    ts   = datetime.today().strftime('%Y%m%d')

    print("\n📊 Building Excel workbook…")
    wb = Workbook()
    wb.remove(wb.active)
    build_summary(wb, all_events, all_attractions, label)
    if all_events:      build_events_sheet(wb, all_events, label)
    if all_attractions: build_attractions_sheet(wb, all_attractions, label)

    xlsx_out = f"events_{safe}_{ts}.xlsx"
    wb.save(xlsx_out)
    print(f"  ✅  Saved: {xlsx_out}")

    print("🗺  Building interactive map…")
    map_out = f"events_{safe}_{ts}_map.html"
    build_map(all_events, all_attractions, label, map_out)
    print(f"  ✅  Saved: {map_out}")

    print(f"\n📋 Summary")
    print(f"    📊 {xlsx_out}")
    print(f"    🗺  {map_out}")
    print(f"    📅 {len(all_events)} events  |  📍 {len(all_attractions)} attractions")
    print(f"\n💡  Filter by City, Category, Date, or Source in Excel row 2.")
    print(f"    Open the .html file in any browser for the interactive map.\n")

if __name__ == "__main__":
    main()
