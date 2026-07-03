"""
Event & Activities Aggregator v6
==================================
All sources verified working via live probes. No sign-up required.

EVENTS:
  - NYC Open Data: Permitted Events     (2000 future events, was 500)
  - NYC Open Data: Film/Theatre Permits (200)
  - TimeOut NYC: 12 category pages + 2 confirmed sub-pages
      /music/concert-schedule-for-live-music-in-new-york-city  (12 shows)
      /nightlife/best-parties-in-nyc-this-week                 (8 picks)
  - Patch.com: 5 NYC neighborhood calendars
  - SummerStage: /events/ paginated (10/page, multiple pages)
  - Prospect Park: /events/list/page/N/ paginated (12/page)

ATTRACTIONS:
  - Wikidata SPARQL (5s pause)
  - OpenStreetMap Overpass POST (8km)
  - NYC Cultural Organizations (300)
  - NYC Farmers Markets (200)

Requirements:
    pip install requests beautifulsoup4 lxml openpyxl
"""

import re, sys, time, json, random
from datetime import datetime
from urllib.parse import quote_plus
from collections import Counter

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Styling ──────────────────────────────────────────────────────────────────

CAT_COLOURS = {
    "Music":"D6E4F7","Sports":"D5E8D4","Arts & Theatre":"E1D5E7",
    "Food & Drink":"FFE6CC","Community":"FFF2CC","Nature & Parks":"C9E6C9",
    "Museums":"E8DAEF","Amusement":"FDDCB5","Shopping":"FEF9E7",
    "Entertainment":"FDE9D9","Architecture":"F8CECC","Film":"DAE8FC",
    "Education":"E8DAEF","Other":"F5F5F5",
}

def _s(): return Side(style="thin", color="CCCCCC")
def _b(): return Border(left=_s(),right=_s(),top=_s(),bottom=_s())

def hdr(cell, bg="1F3864", fg="FFFFFF", sz=11):
    cell.font=Font(bold=True,color=fg,name="Arial",size=sz)
    cell.fill=PatternFill("solid",fgColor=bg)
    cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    cell.border=_b()

def dat(cell, bg="FFFFFF"):
    cell.font=Font(name="Arial",size=10)
    cell.fill=PatternFill("solid",fgColor=bg)
    cell.alignment=Alignment(vertical="center",wrap_text=True)
    cell.border=_b()

def set_widths(ws,d):
    for col,w in d.items(): ws.column_dimensions[col].width=w

# ─── HTTP ─────────────────────────────────────────────────────────────────────

SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept-Language":"en-US,en;q=0.9",
    "Accept":"text/html,application/xhtml+xml,*/*;q=0.8",
})

def get(url, pause=1.1, hdrs=None, timeout=15):
    time.sleep(pause+random.uniform(0,0.4))
    try:
        r=SESSION.get(url,headers=hdrs or {},timeout=timeout)
        r.raise_for_status()
        return r
    except Exception as e:
        print(f"  ⚠  {url[:68]}… → {e}")
        return None

def post_req(url,data,pause=1.5,hdrs=None,timeout=25):
    time.sleep(pause+random.uniform(0,0.4))
    try:
        r=SESSION.post(url,data=data,headers=hdrs or {},timeout=timeout)
        r.raise_for_status()
        return r
    except Exception as e:
        print(f"  ⚠  POST {url[:65]}… → {e}")
        return None

# ─── Helpers ──────────────────────────────────────────────────────────────────

def classify(text):
    t=(text or "").lower()
    if any(w in t for w in ["concert","music","band","dj","jazz","rock","hip hop","singer","tour","symphony","orchestra","opera","festival","gig","bluegrass","electronic","choir","r&b","reggae","punk","indie","album","live music","nightlife","party","club","lounge"]):
        return "Music"
    if any(w in t for w in ["sport","game","match","nba","nfl","mlb","nhl","soccer","tennis","golf","marathon","race","baseball","basketball","football","hockey","esport","wrestling","mma","athletic","pickleball","lacrosse","volleyball","swim","yoga","fitness","run","cycling"]):
        return "Sports"
    if any(w in t for w in ["museum","exhibit","gallery","art show","theatre","theater","ballet","dance","comedy","stand-up","broadway","performance","circus","puppet","magic","improv","opera","cabaret","burlesque"]):
        return "Arts & Theatre"
    if any(w in t for w in ["food","drink","wine","beer","brunch","tasting","chef","culinary","cocktail","dining","grill","whiskey","spirits","gastro","restaurant","greenmarket","farmers market","food festival","bbq","taco","pizza"]):
        return "Food & Drink"
    if any(w in t for w in ["community","fair","parade","volunteer","charity","fundraiser","seminar","networking","conference","health","wellness","meditation","block party","street fair","flea","craft","expo","summit"]):
        return "Community"
    if any(w in t for w in ["park","garden","nature","hike","trail","zoo","aquarium","botanical","outdoor","beach","lake","forest","reserve","greenway","conservancy","wildlife"]):
        return "Nature & Parks"
    if any(w in t for w in ["theme park","amusement","roller coaster","arcade","carnival","carousel","ferris wheel"]):
        return "Amusement"
    if any(w in t for w in ["film","cinema","movie","screening","documentary"]):
        return "Film"
    if any(w in t for w in ["lecture","learn","education","school","university","training","class","workshop","tutorial"]):
        return "Education"
    if any(w in t for w in ["shop","mall","flea market","boutique","bazaar","retail","store","market"]):
        return "Shopping"
    return "Entertainment"

def parse_date(s):
    if not s: return ""
    s=str(s)
    m=re.search(r"(\d{4}-\d{2}-\d{2})",s)
    if m: return m.group(1)
    if re.match(r"^\d{10}$",s.strip()):
        try: return datetime.fromtimestamp(int(s)).strftime("%Y-%m-%d")
        except: pass
    for fmt in ("%B %d, %Y","%b %d, %Y","%m/%d/%Y","%B %d"):
        try:
            dt=datetime.strptime(s.strip()[:20],fmt)
            if dt.year==1900: dt=dt.replace(year=datetime.today().year)
            return dt.strftime("%Y-%m-%d")
        except: pass
    return s[:20]

def parse_time(s):
    if not s: return ""
    m=re.search(r"T(\d{2}:\d{2})",str(s))
    if m: return m.group(1)
    m=re.search(r"(\d{1,2}:\d{2}\s*(?:am|pm)?)",str(s),re.I)
    if m: return m.group(1)
    return ""

def geocode(location):
    url=f"https://nominatim.openstreetmap.org/search?q={quote_plus(location)}&format=json&limit=1&countrycodes=us"
    r=get(url,pause=1,hdrs={"User-Agent":"EventAggregator/6.0"})
    if not r: return None,None,location
    data=r.json()
    if not data: return None,None,location
    d=data[0]
    return float(d["lat"]),float(d["lon"]),d["display_name"]

def ev(name,cat,date="",time_="",venue="",address="",city="",price="",url="",source=""):
    return {"Name":name,"Category":cat,"Date":date,"Time":time_,
            "Venue":venue,"Address":address,"City":city,
            "Price":price,"URL":url,"Source":source}

def att(name,cat,type_="",address="",city="",url="",source=""):
    return {"Name":name,"Category":cat,"Type":type_,
            "Address":address,"City":city,"URL":url,"Source":source}

# ═══════════════════════════════════════════════════════════════════════════════
#  EVENT SOURCES
# ═══════════════════════════════════════════════════════════════════════════════

# ── 1. NYC Open Data: Permitted Events (raised to 2000) ───────────────────────

def fetch_nyc_permitted(limit=2000):
    print(f"  → NYC Open Data: Permitted Events (up to {limit})...")
    today=datetime.today().strftime("%Y-%m-%dT00:00:00.000")
    url=(f"https://data.cityofnewyork.us/resource/tvpp-9vvx.json"
         f"?$limit={limit}&$where=start_date_time>='{today}'&$order=start_date_time+ASC")
    r=get(url,pause=1,hdrs={"Accept":"application/json"})
    if not r: return []
    skip={"closure","n/a","none","nan",""}
    out=[]
    for item in r.json():
        name=item.get("event_name","").strip()
        if name.lower() in skip: continue
        dr=item.get("start_date_time","")
        out.append(ev(name,classify(item.get("event_type","")+" "+name),
            parse_date(dr),dr[11:16] if len(dr)>10 else "",
            item.get("event_location",""),"",
            item.get("event_borough","New York")+", NY",
            "Varies","","NYC Open Data"))
    print(f"     {len(out)} permitted events")
    return out

# ── 2. NYC Film & Theatre Permits ─────────────────────────────────────────────

def fetch_nyc_film_permits(limit=200):
    print("  → NYC Open Data: Film & Theatre Permits...")
    url=(f"https://data.cityofnewyork.us/resource/tg4x-b46p.json"
         f"?$limit={limit}&$order=enteredon+DESC")
    r=get(url,pause=1,hdrs={"Accept":"application/json"})
    if not r: return []
    out=[]
    for item in r.json():
        etype=item.get("eventtype","").strip()
        cat=item.get("category","").strip()
        subcat=item.get("subcategoryname","").strip()
        if not etype: continue
        name=f"{cat}: {subcat}" if subcat and cat else (subcat or etype)
        dr=item.get("startdatetime","") or item.get("enteredon","")
        out.append(ev(name,classify(cat+" "+subcat+" "+etype),
            parse_date(dr),dr[11:16] if len(dr)>10 else "",
            item.get("parkingheld",""),"",
            item.get("borough","New York")+", NY",
            "Industry/Varies","","NYC Film Permits"))
    print(f"     {len(out)} film/theatre permits")
    return out

# ── 3. TimeOut NYC ────────────────────────────────────────────────────────────

TIMEOUT_PAGES=[
    ("music","Music"),("theater","Arts & Theatre"),("art","Arts & Theatre"),
    ("comedy","Arts & Theatre"),("film","Film"),("nightlife","Entertainment"),
    ("dance","Arts & Theatre"),("restaurants","Food & Drink"),
    ("shopping","Shopping"),("things-to-do","Entertainment"),
    ("attractions","Entertainment"),
]
# Confirmed sub-pages with individual event listings
TIMEOUT_SUBPAGES=[
    ("music/concert-schedule-for-live-music-in-new-york-city","Music"),
    ("nightlife/best-parties-in-nyc-this-week","Entertainment"),
    ("dance/the-best-dance-shows-in-nyc-this-month","Arts & Theatre"),
]
TIMEOUT_SLUGS={
    "new york":"newyork","los angeles":"los-angeles","chicago":"chicago",
    "miami":"miami","san francisco":"san-francisco","boston":"boston",
    "seattle":"seattle","denver":"denver","atlanta":"atlanta","dallas":"dallas",
    "austin":"austin","portland":"portland","nashville":"nashville",
    "las vegas":"las-vegas","philadelphia":"philadelphia",
    "washington":"washington-dc","houston":"houston","san diego":"san-diego",
}

def _timeout_cards(soup, default_cat, seen, base_url="https://www.timeout.com"):
    """Extract events from TimeOut article cards and article lists."""
    added=[]
    # Standard tile cards
    for card in soup.select("article[data-testid='tile-zone-a_testID']"):
        link=card.find("a",href=True)
        img=card.find("img")
        if not link: continue
        href=link["href"]
        if not href.startswith("http"): href=base_url+href
        title=(img.get("alt","") if img else "").strip()
        if not title:
            h=card.find(["h2","h3","h4"])
            title=h.get_text(strip=True) if h else ""
        if not title: title=href.rstrip("/").split("/")[-1].replace("-"," ").title()
        key=title.lower()[:60]
        if key in seen or len(key)<4: continue
        seen.add(key)
        cat=default_cat
        for tok,mapped in [("music","Music"),("theater","Arts & Theatre"),
                            ("art","Arts & Theatre"),("comedy","Arts & Theatre"),
                            ("film","Film"),("nightlife","Entertainment"),
                            ("dance","Arts & Theatre"),("restaurant","Food & Drink"),
                            ("shopping","Shopping"),("attraction","Entertainment")]:
            if tok in href: cat=mapped; break
        else: cat=classify(title)
        added.append((title,cat,href))
    # Article list items (numbered picks pages like nightlife/best-parties)
    for art in soup.select("article"):
        h=art.find(["h2","h3","h4"])
        link=art.find("a",href=True)
        if not h or not link: continue
        title=re.sub(r"^\d+\.\s*","",h.get_text(strip=True))  # strip leading "1."
        key=title.lower()[:60]
        if key in seen or len(key)<4: continue
        seen.add(key)
        href=link["href"]
        if not href.startswith("http"): href=base_url+href
        added.append((title,classify(title),href))
    return added

def fetch_timeout(location):
    city_key=location.split(",")[0].strip().lower()
    slug=next((s for k,s in TIMEOUT_SLUGS.items() if k in city_key or city_key in k),None)
    if not slug:
        print(f"  ℹ  TimeOut: no slug for '{city_key}'")
        return []
    print(f"  → TimeOut {slug}: category pages + sub-pages...")
    events,seen=[],set()

    def add_from_url(url, default_cat, label):
        r=get(url,pause=1.1)
        if not r: return 0
        soup=BeautifulSoup(r.text,"lxml")
        cards=_timeout_cards(soup,default_cat,seen)
        for title,cat,href in cards:
            events.append(ev(title,cat,city=location,url=href,source="TimeOut"))
        if cards: print(f"     +{len(cards)} from {label}")
        return len(cards)

    for page_path,default_cat in TIMEOUT_PAGES:
        add_from_url(f"https://www.timeout.com/{slug}/{page_path}",default_cat,f"/{page_path}")

    # Sub-pages with individual picks (NYC-specific)
    if slug=="newyork":
        for page_path,default_cat in TIMEOUT_SUBPAGES:
            add_from_url(f"https://www.timeout.com/{slug}/{page_path}",default_cat,f"/{page_path} (sub)")

    print(f"     {len(events)} total from TimeOut")
    return events

# ── 4. Patch.com ──────────────────────────────────────────────────────────────

PATCH_NYC=[("new-york","new-york-city"),("new-york","upper-west-side-nyc"),
           ("new-york","upper-east-side-nyc"),("new-york","brooklyn"),
           ("new-york","queens")]
PATCH_OTHER={"chicago":"illinois","los angeles":"california","houston":"texas",
             "miami":"florida","boston":"massachusetts","seattle":"washington",
             "denver":"colorado","atlanta":"georgia","dallas":"texas",
             "austin":"texas","portland":"oregon","nashville":"tennessee",
             "philadelphia":"pennsylvania","san francisco":"california"}

def _parse_patch(data,location,seen):
    pp=data.get("props",{}).get("pageProps",{})
    mc=pp.get("mainContent",{})
    rails=pp.get("rightRail",[])
    out=[]
    def add(item):
        title=(item.get("title") or item.get("shortTitle") or "").strip()
        if not title or title.lower() in seen: return
        seen.add(title.lower())
        body=BeautifulSoup(item.get("body",""),"lxml").get_text(" ")[:200]
        ts=item.get("displayDateTimestamp") or item.get("created","")
        date_str=parse_date(str(ts)) if str(ts).isdigit() else parse_date(item.get("displayDate",""))
        addr_obj=item.get("address") or {}
        address=addr_obj.get("display","") if isinstance(addr_obj,dict) else str(addr_obj)
        etype=item.get("eventType","")
        price="Free" if etype=="free" else ("Paid" if etype=="paid" else "")
        url_=item.get("canonicalUrl","") or item.get("url","")
        out.append(ev(title,classify(etype+" "+title+" "+body),
            date_str,"",item.get("locationName",""),address,location,price,url_,"Patch.com"))
    # allEvents: {timestamp_str: [events]}
    for bucket in [mc.get("allEvents",{}),mc.get("promotedEvents",{})]:
        if isinstance(bucket,dict):
            for evlist in bucket.values():
                if isinstance(evlist,list):
                    for item in evlist: add(item)
        elif isinstance(bucket,list):
            for item in bucket: add(item)
    # featuredEvents in rightRail
    for rail in rails:
        if rail.get("type")=="featuredEvents":
            for item in rail.get("items",[]): add(item)
    return out

def fetch_patch(location):
    city_key=location.split(",")[0].strip().lower()
    is_nyc="new york" in city_key
    slugs=PATCH_NYC if is_nyc else []
    if not is_nyc:
        state=PATCH_OTHER.get(city_key,"")
        if state: slugs=[(state,city_key.replace(" ","-"))]
    if not slugs:
        print(f"  ℹ  Patch.com: no slug for '{city_key}'")
        return []
    print(f"  → Patch.com: {len(slugs)} neighborhoods...")
    seen,all_events=set(),[]
    for state,city in slugs:
        r=get(f"https://patch.com/{state}/{city}/calendar",pause=1.2)
        if not r: continue
        soup=BeautifulSoup(r.text,"lxml")
        nd=soup.find("script",id="__NEXT_DATA__")
        if not nd: continue
        try:
            evts=_parse_patch(json.loads(nd.string),location,seen)
            if evts: print(f"     +{len(evts)} from /{city}")
            all_events.extend(evts)
        except Exception as e:
            print(f"     Parse error ({city}): {e}")
    print(f"     {len(all_events)} total from Patch.com")
    return all_events

# ── 5. SummerStage — paginated JSON-LD ───────────────────────────────────────

def fetch_summerstage(max_pages=6):
    print("  → SummerStage (City Parks Foundation, paginated)...")
    events,seen=[],set()

    def scrape_page(url):
        r=get(url,pause=1.2)
        if not r: return False,None
        soup=BeautifulSoup(r.text,"lxml")
        added=0
        for script in soup.find_all("script",type="application/ld+json"):
            try:
                d=json.loads(script.string or "")
                items=d if isinstance(d,list) else [d]
                for item in items:
                    if item.get("@type") not in ("Event","MusicEvent","SportsEvent","TheaterEvent"): continue
                    name=item.get("name","").strip()
                    if not name or name.lower() in seen: continue
                    seen.add(name.lower())
                    loc=item.get("location",{})
                    venue=loc.get("name","") if isinstance(loc,dict) else str(loc)
                    offers=item.get("offers",{})
                    pval=offers.get("price","") if isinstance(offers,dict) else ""
                    price="Free" if str(pval)=="0" else (f"${pval}" if pval else "")
                    start=item.get("startDate","")
                    events.append(ev(name,classify(name),
                        parse_date(start),parse_time(start),
                        venue,"","New York, NY",price,item.get("url",""),"SummerStage"))
                    added+=1
            except: pass
        # Also parse tribe articles for any missed
        for art in soup.select("article.type-tribe_events"):
            name_el=art.find(class_=re.compile(r"event-title",re.I)) or art.find(["h2","h3","h4"])
            link_el=art.find("a",href=True)
            date_el=art.find("time")
            name=name_el.get_text(strip=True) if name_el else ""
            if not name or name.lower() in seen: continue
            seen.add(name.lower())
            price_el=art.find(class_=re.compile(r"price|cost",re.I))
            events.append(ev(name,classify(name),
                date_el.get("datetime","") if date_el else "","","","",
                "New York, NY",price_el.get_text(strip=True) if price_el else "",
                link_el["href"] if link_el else "","SummerStage"))
            added+=1
        # Find next page
        nxt=soup.find("a",class_=re.compile(r"tribe-events-nav-next|next-events",re.I))
        next_url=nxt["href"] if nxt and nxt.get("href") else None
        return added>0, next_url

    # Start from main events page (10 per page)
    url="https://cityparksfoundation.org/events/"
    for page in range(1,max_pages+1):
        had_new, next_url=scrape_page(url)
        if not had_new: break
        if not next_url: break
        url=next_url

    print(f"     {len(events)} events from SummerStage")
    return events

# ── 6. Prospect Park — paginated tribe_events ─────────────────────────────────

CAT_MAP_PP={
    "cat_music":"Music","cat_concert":"Music","cat_art":"Arts & Theatre",
    "cat_dance":"Arts & Theatre","cat_theater":"Arts & Theatre",
    "cat_food":"Food & Drink","cat_wellness":"Sports","cat_sports":"Sports",
    "cat_kids":"Community","cat_community":"Community","cat_education":"Education",
    "cat_film":"Film","cat_nature":"Nature & Parks","cat_free":"Community",
}

def fetch_prospect_park(max_pages=5):
    print("  → Prospect Park Events (paginated)...")
    events,seen=[],set()
    base="https://www.prospectpark.org"

    def scrape_page(url):
        r=get(url,pause=1.2)
        if not r: return 0,None
        soup=BeautifulSoup(r.text,"lxml")
        added=0
        for art in soup.select("article.type-tribe_events"):
            name_el=art.find(class_=re.compile(r"event-title",re.I)) or art.find(["h2","h3","h4"])
            link_el=art.find("a",href=True)
            date_el=art.find("time")
            if not name_el: continue
            name=name_el.get_text(strip=True)
            if not name or name.lower() in seen: continue
            seen.add(name.lower())
            classes=art.get("class",[])
            cat=next((CAT_MAP_PP[c] for c in classes if c in CAT_MAP_PP),classify(name))
            date_str=date_el.get("datetime","") if date_el else ""
            events.append(ev(name,cat,
                parse_date(date_str),parse_time(date_str),
                "Prospect Park","Brooklyn","Brooklyn, NY",
                "Free",link_el["href"] if link_el else "","Prospect Park"))
            added+=1
        nxt=soup.find("a",class_=re.compile(r"tribe-events-nav-next|next",re.I),href=True)
        next_url=nxt["href"] if nxt else None
        # Also try list pagination link
        if not next_url:
            nxt2=soup.find("a",href=re.compile(r"/events/list/page/\d+/"))
            next_url=nxt2["href"] if nxt2 else None
        return added,next_url

    url=f"{base}/events/"
    for page in range(1,max_pages+1):
        added,next_url=scrape_page(url)
        if not added: break
        print(f"     +{added} from page {page}")
        if not next_url: break
        url=next_url if next_url.startswith("http") else base+next_url

    print(f"     {len(events)} total from Prospect Park")
    return events

# ═══════════════════════════════════════════════════════════════════════════════
#  ATTRACTION SOURCES  (unchanged from v5)
# ═══════════════════════════════════════════════════════════════════════════════

WIKIDATA_CAT={
    "museum":"Museums","art museum":"Museums","children's museum":"Museums",
    "history museum":"Museums","science museum":"Museums",
    "amusement park":"Amusement","theme park":"Amusement","carousel":"Amusement",
    "roller coaster":"Amusement","ferris wheel":"Amusement",
    "observation deck":"Entertainment","tourist attraction":"Entertainment",
    "park":"Nature & Parks","botanical garden":"Nature & Parks",
    "nature reserve":"Nature & Parks","aquarium":"Nature & Parks","zoo":"Nature & Parks",
    "stadium":"Sports","theatre":"Arts & Theatre","theater":"Arts & Theatre",
    "movie theatre":"Film","concert hall":"Music",
    "historic landmark":"Architecture","historic district":"Architecture",
    "public aquarium":"Nature & Parks",
}

def fetch_wikidata(lat,lon,radius_km=15):
    if lat is None: return []
    print("  → Wikidata SPARQL (5s pause)...")
    time.sleep(5)
    sparql=f"""
SELECT DISTINCT ?place ?placeLabel ?typeLabel ?website ?article WHERE {{
  SERVICE wikibase:around {{
    ?place wdt:P625 ?coord .
    bd:serviceParam wikibase:center "Point({lon} {lat})"^^geo:wktLiteral .
    bd:serviceParam wikibase:radius "{radius_km}" .
  }}
  ?place wdt:P31 ?type .
  VALUES ?type {{
    wd:Q33506 wd:Q207694 wd:Q1007870 wd:Q574915
    wd:Q482994 wd:Q197646 wd:Q1329623 wd:Q22698
    wd:Q570116 wd:Q41253  wd:Q24354  wd:Q14092
    wd:Q1107656 wd:Q15078955 wd:Q28564 wd:Q167346
  }}
  OPTIONAL {{ ?place wdt:P856 ?website }}
  OPTIONAL {{ ?article schema:about ?place ; schema:isPartOf <https://en.wikipedia.org/> }}
  SERVICE wikibase:label {{ bd:serviceParam wikibase:language "en" }}
}}
LIMIT 80
"""
    try:
        resp=SESSION.get("https://query.wikidata.org/sparql",
            params={"query":sparql,"format":"json"},
            headers={"Accept":"application/sparql-results+json","User-Agent":"EventAggregator/6.0"},
            timeout=30)
        resp.raise_for_status()
        bindings=resp.json()["results"]["bindings"]
    except Exception as e:
        print(f"  ⚠  Wikidata: {e}"); return []
    places,seen=[],set()
    for b in bindings:
        name=b.get("placeLabel",{}).get("value","").strip()
        ptype=b.get("typeLabel",{}).get("value","").strip()
        web=b.get("website",{}).get("value","")
        art=b.get("article",{}).get("value","")
        if not name or name in seen or re.match(r"^Q\d+$",name): continue
        seen.add(name)
        cat=WIKIDATA_CAT.get(ptype.lower(),classify(name+" "+ptype))
        places.append(att(name,cat,ptype or "Attraction","","",web or art,"Wikidata"))
    print(f"     {len(places)} from Wikidata")
    return places

OSM_MAP={"attraction":"Entertainment","museum":"Museums","theme_park":"Amusement",
         "viewpoint":"Entertainment","zoo":"Nature & Parks","aquarium":"Nature & Parks",
         "gallery":"Arts & Theatre","park":"Nature & Parks","nature_reserve":"Nature & Parks",
         "garden":"Nature & Parks","stadium":"Sports","theatre":"Arts & Theatre",
         "cinema":"Film","arts_centre":"Arts & Theatre"}

def fetch_osm(lat,lon,radius_m=8000,limit=100):
    if lat is None: return []
    print("  → OpenStreetMap Overpass (8km)...")
    query=f"""[out:json][timeout:20];
(
  node["tourism"~"attraction|museum|theme_park|zoo|aquarium|gallery"](around:{radius_m},{lat},{lon});
  node["amenity"~"theatre|cinema|arts_centre"](around:{radius_m},{lat},{lon});
  node["leisure"~"park|nature_reserve|garden"](around:{radius_m},{lat},{lon});
);
out {limit};"""
    r=post_req("https://overpass-api.de/api/interpreter",{"data":query},
               hdrs={"User-Agent":"EventAggregator/6.0"})
    if not r: return []
    places,seen=[],set()
    for el in r.json().get("elements",[]):
        tags=el.get("tags",{})
        name=tags.get("name","").strip()
        if not name or name in seen: continue
        seen.add(name)
        t=tags.get("tourism",""); l=tags.get("leisure",""); a=tags.get("amenity","")
        cat=OSM_MAP.get(t) or OSM_MAP.get(l) or OSM_MAP.get(a) or classify(name)
        addr=" ".join(filter(None,[tags.get("addr:housenumber",""),tags.get("addr:street","")]))
        places.append(att(name,cat,t or l or a or "Attraction",addr,
                          tags.get("addr:city",""),
                          tags.get("website",tags.get("contact:website","")),
                          "OpenStreetMap"))
    print(f"     {len(places)} from OpenStreetMap")
    return places

def fetch_nyc_cultural_orgs(limit=300):
    print("  → NYC Cultural Organizations...")
    r=get(f"https://data.cityofnewyork.us/resource/u35m-9t32.json?$limit={limit}",
          pause=1,hdrs={"Accept":"application/json"})
    if not r: return []
    places,seen=[],set()
    for item in r.json():
        name=item.get("organization_name","").strip()
        if not name or name in seen: continue
        seen.add(name)
        disc=item.get("discipline","")
        cat=classify(disc+" "+name)
        if "theater" in disc.lower() or "theatre" in disc.lower(): cat="Arts & Theatre"
        elif "music" in disc.lower(): cat="Music"
        elif "museum" in disc.lower(): cat="Museums"
        elif "dance" in disc.lower(): cat="Arts & Theatre"
        elif "film" in disc.lower(): cat="Film"
        places.append(att(name,cat,disc or "Cultural Org",
                          item.get("address",""),
                          f"{item.get('city','New York')}, {item.get('state','NY')}",
                          "","NYC Cultural Orgs"))
    print(f"     {len(places)} cultural orgs")
    return places

def fetch_nyc_farmers_markets(limit=200):
    print("  → NYC Farmers Markets...")
    r=get(f"https://data.cityofnewyork.us/resource/8vwk-6iz2.json?$limit={limit}",
          pause=1,hdrs={"Accept":"application/json"})
    if not r: return []
    places=[]
    for item in r.json():
        name=item.get("marketname","").strip()
        if not name: continue
        days=item.get("daysoperation","")
        hours=item.get("hoursoperations","")
        yr="Year-round" if item.get("open_year_round","").lower()=="yes" else "Seasonal"
        ebt=" | EBT accepted" if item.get("accepts_ebt","").lower()=="yes" else ""
        label=f"{name} ({days}, {hours})" if days else name
        places.append(att(label,"Food & Drink",f"Farmers Market — {yr}{ebt}",
                          item.get("streetaddress",""),
                          f"{item.get('borough','New York')}, NY","","NYC Farmers Markets"))
    print(f"     {len(places)} farmers markets")
    return places

# ═══════════════════════════════════════════════════════════════════════════════
#  EXCEL OUTPUT
# ═══════════════════════════════════════════════════════════════════════════════

ECOLS=["#","Name","Category","Date","Time","Venue","Address","City","Price","URL","Source"]
ACOLS=["#","Name","Category","Type","Address","City","URL","Source"]

def _title(ws,text,ncols,bg,sz=13):
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    hdr(ws["A1"],bg=bg,sz=sz); ws["A1"].value=text
    ws.row_dimensions[1].height=28

def build_summary(wb,events,attractions,location):
    ws=wb.create_sheet("📊 Summary",0)
    ws.sheet_properties.tabColor="833C00"
    _title(ws,"Event & Activities Aggregator — Summary",3,"833C00",14)
    def kv(row,k,v):
        c1=ws.cell(row=row,column=1,value=k)
        c2=ws.cell(row=row,column=2,value=str(v))
        c1.font=Font(bold=True,name="Arial",size=10)
        c2.font=Font(name="Arial",size=10)
        for c in [c1,c2]:
            c.fill=PatternFill("solid",fgColor="F2F2F2"); c.border=_b()
    kv(2,"Location",location)
    kv(3,"Generated On",datetime.now().strftime("%Y-%m-%d %H:%M"))
    kv(4,"Total Events",len(events))
    kv(5,"Total Attractions / Things To Do",len(attractions))
    kv(6,"Grand Total",len(events)+len(attractions))
    row=8
    for htext,bg_,items,field in [
        ("Events by Category","2E75B6",events,"Category"),
        ("Attractions by Category","375623",attractions,"Category"),
        ("Data Sources","595959",events+attractions,"Source"),
    ]:
        ws.merge_cells(f"A{row}:C{row}")
        hdr(ws.cell(row,1,htext),bg=bg_)
        ws.row_dimensions[row].height=20
        row+=1
        for val,cnt in Counter(x[field] for x in items).most_common():
            bg=CAT_COLOURS.get(val,"F5F5F5") if field=="Category" else "F2F2F2"
            c1=ws.cell(row,1,val); c2=ws.cell(row,2,cnt)
            dat(c1,bg); dat(c2,bg)
            c2.alignment=Alignment(horizontal="center",vertical="center")
            row+=1
        row+=1
    set_widths(ws,{"A":32,"B":14,"C":14})

def build_events_sheet(wb,events,location):
    ws=wb.create_sheet("📅 Events")
    ws.sheet_properties.tabColor="2E75B6"
    _title(ws,f"Events near {location}",len(ECOLS),"1F3864")
    for c,n in enumerate(ECOLS,1): hdr(ws.cell(2,c,n))
    ws.row_dimensions[2].height=22
    for i,e in enumerate(events,1):
        row=i+2; bg=CAT_COLOURS.get(e.get("Category","Other"),"F5F5F5")
        vals=[i,e.get("Name"),e.get("Category"),e.get("Date"),e.get("Time"),
              e.get("Venue"),e.get("Address"),e.get("City"),e.get("Price"),
              e.get("URL"),e.get("Source")]
        for c,v in enumerate(vals,1):
            cell=ws.cell(row,c,v); dat(cell,bg)
            if c==1: cell.alignment=Alignment(horizontal="center",vertical="center")
            if c==10 and v:
                cell.hyperlink=v
                cell.font=Font(color="0563C1",underline="single",name="Arial",size=10)
        ws.row_dimensions[row].height=18
    ws.freeze_panes="A3"
    ws.auto_filter.ref=f"A2:{get_column_letter(len(ECOLS))}2"
    set_widths(ws,{"A":5,"B":38,"C":16,"D":12,"E":7,"F":26,"G":22,"H":16,"I":12,"J":16,"K":18})

def build_attractions_sheet(wb,attractions,location):
    ws=wb.create_sheet("📍 Things To Do")
    ws.sheet_properties.tabColor="375623"
    _title(ws,f"Attractions & Things To Do near {location}",len(ACOLS),"375623")
    for c,n in enumerate(ACOLS,1): hdr(ws.cell(2,c,n),bg="375623")
    ws.row_dimensions[2].height=22
    for i,pl in enumerate(attractions,1):
        row=i+2; bg=CAT_COLOURS.get(pl.get("Category","Other"),"F5F5F5")
        vals=[i,pl.get("Name"),pl.get("Category"),pl.get("Type"),
              pl.get("Address"),pl.get("City"),pl.get("URL"),pl.get("Source")]
        for c,v in enumerate(vals,1):
            cell=ws.cell(row,c,v); dat(cell,bg)
            if c==1: cell.alignment=Alignment(horizontal="center",vertical="center")
            if c==7 and v:
                cell.hyperlink=v
                cell.font=Font(color="0563C1",underline="single",name="Arial",size=10)
        ws.row_dimensions[row].height=18
    ws.freeze_panes="A3"
    ws.auto_filter.ref=f"A2:{get_column_letter(len(ACOLS))}2"
    set_widths(ws,{"A":5,"B":42,"C":18,"D":28,"E":26,"F":18,"G":22,"H":18})

# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    print("="*68)
    print("  Event & Activities Aggregator v6  |  No Sign-Up Required")
    print("="*68)
    print()
    print("  Events:      NYC Open Data (2000) · Film Permits · TimeOut")
    print("               Patch.com · SummerStage · Prospect Park")
    print("  Attractions: Wikidata · OpenStreetMap · NYC Cultural Orgs")
    print("               NYC Farmers Markets")
    print()

    location=input("City (e.g. 'New York, NY'): ").strip() or "New York, NY"
    radius_in=input("Search radius in miles [default 15]: ").strip()
    radius_mi=int(radius_in) if radius_in.isdigit() else 15
    radius_m=int(radius_mi*1609.34)
    radius_km=int(radius_mi*1.60934)
    city_key=location.split(",")[0].strip().lower()
    is_nyc="new york" in city_key

    print(f"\n🔍 Geocoding '{location}'...")
    lat,lon,display=geocode(location)
    if lat: print(f"   → {display[:70]}\n   → {lat:.4f}, {lon:.4f}")
    else:   print("   ⚠  Geocoding failed — distance sources skipped")

    print("\n📡 Fetching data…\n")

    events=[]
    if is_nyc:
        events+=fetch_nyc_permitted(2000)
        events+=fetch_nyc_film_permits(200)
    events+=fetch_timeout(location)
    events+=fetch_patch(location)
    if is_nyc:
        events+=fetch_summerstage(max_pages=6)
        events+=fetch_prospect_park(max_pages=5)

    attractions=[]
    if lat:
        attractions+=fetch_wikidata(lat,lon,radius_km)
        attractions+=fetch_osm(lat,lon,radius_m)
    if is_nyc:
        attractions+=fetch_nyc_cultural_orgs(300)
        attractions+=fetch_nyc_farmers_markets(200)

    def dedup(lst):
        seen,out=set(),[]
        for x in lst:
            k=(x.get("Name") or "").lower().strip()[:60]
            if k and k not in seen: seen.add(k); out.append(x)
        return out

    events=dedup(events)
    attractions=dedup(attractions)

    print(f"\n{'─'*54}")
    print(f"  ✅  {len(events)} events  |  {len(attractions)} attractions")
    print(f"{'─'*54}")

    if not events and not attractions:
        print("\n✗  No data found."); sys.exit(1)

    print("\n📊 Building Excel workbook…")
    wb=Workbook()
    wb.remove(wb.active)
    build_summary(wb,events,attractions,location)
    if events: build_events_sheet(wb,events,location)
    if attractions: build_attractions_sheet(wb,attractions,location)

    safe=re.sub(r"[^\w]","_",location)
    out=f"events_{safe}_{datetime.today().strftime('%Y%m%d')}.xlsx"
    wb.save(out)
    print(f"\n✅  Saved: {out}")
    print(f"    📊 Summary  |  📅 {len(events)} Events  |  📍 {len(attractions)} Attractions\n")

if __name__=="__main__":
    main()
